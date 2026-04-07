# backend/modules/impresion.py
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import filedialog
import os
import tempfile
from datetime import datetime
import win32con  # Agregar esta línea si está disponible

class VistaPreviaImpresion:
    def __init__(self, parent, titulo, datos, columnas=None, empresa=None):
        self.parent = parent
        self.titulo = titulo
        self.datos = datos
        self.columnas = columnas
        self.empresa = empresa
        self._logo_img = None
        self.crear_vista_previa()
    
    def crear_vista_previa(self):
        """Crea ventana de vista previa de impresión (formato empresarial si hay empresa)."""
        self.ventana = tk.Toplevel(self.parent)
        self.ventana.title(f"Vista Previa - {self.titulo}")
        self.ventana.geometry("850x620")
        
        toolbar = tk.Frame(self.ventana, bg='#1e3a5f', height=44)
        toolbar.pack(fill=tk.X)
        tk.Button(toolbar, text="Vista previa / Imprimir", command=self.imprimir,
                 bg='#2563eb', fg='white', relief=tk.FLAT, padx=14, pady=6, cursor='hand2',
                 font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT, padx=8, pady=6)
        tk.Button(toolbar, text="Guardar como PDF", command=self.guardar_como_pdf,
                 bg='#475569', fg='white', relief=tk.FLAT, padx=12, pady=6, cursor='hand2').pack(side=tk.LEFT, padx=2)
        tk.Button(toolbar, text="Cerrar", command=self.ventana.destroy,
                 bg='#64748b', fg='white', relief=tk.FLAT, padx=12, pady=6, cursor='hand2').pack(side=tk.RIGHT, padx=8, pady=6)
        
        self.frame_vista = tk.Frame(self.ventana, bg='white', relief=tk.SUNKEN, bd=2)
        self.frame_vista.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.mostrar_vista_previa()
    
    def mostrar_vista_previa(self):
        """Muestra el contenido en vista previa con formato empresarial."""
        canvas = tk.Canvas(self.frame_vista, bg='white')
        canvas.pack(fill=tk.BOTH, expand=True)
        y = 24
        if self.empresa:
            # Logo SSEPI (si existe)
            try:
                logo_path = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "assets", "ssepi_logo.png"))
                if os.path.exists(logo_path):
                    self._logo_img = tk.PhotoImage(file=logo_path)
                    # escalar a ~32px de alto
                    h = max(1, int(self._logo_img.height() / 32))
                    w = max(1, int(self._logo_img.width() / 32))
                    self._logo_img = self._logo_img.subsample(w, h)
                    canvas.create_image(60, y+6, image=self._logo_img, anchor='w')
            except Exception:
                pass
            canvas.create_text(400, y, text=self.empresa, font=('Arial', 12, 'bold'), fill='#1e3a5f')
            y += 28
            canvas.create_line(80, y, 720, y, fill='#cbd5e1', width=1)
            y += 20
        canvas.create_text(400, y, text=self.titulo, font=('Arial', 14, 'bold'), fill='#000080')
        y += 28
        canvas.create_text(400, y, text=f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", font=('Arial', 9), fill='#64748b')
        y += 24
        if isinstance(self.datos, list) and self.columnas:
            self.mostrar_tabla(canvas, y)
        else:
            self.mostrar_texto(canvas, y)
    
    def mostrar_tabla(self, canvas, y):
        """Muestra datos en formato tabla"""
        # Encabezados
        x = 50
        for col in self.columnas:
            canvas.create_rectangle(x-2, y-2, x+120, y+20, outline='black')
            canvas.create_text(x+58, y+8, text=col, font=('Arial', 9, 'bold'))
            x += 120
        
        y += 25
        
        # Datos
        for fila in self.datos[:20]:  # Mostrar primeras 20 filas
            x = 50
            for valor in fila:
                canvas.create_rectangle(x-2, y-2, x+120, y+20, outline='#E0E0E0')
                canvas.create_text(x+58, y+8, text=str(valor), font=('Arial', 8))
                x += 120
            y += 20
        
        if len(self.datos) > 20:
            canvas.create_text(400, y+20, text=f"... y {len(self.datos)-20} filas más",
                              font=('Arial', 9, 'italic'))
    
    def mostrar_texto(self, canvas, y):
        """Muestra datos en formato texto (str o dict; si dict anidado, formatea por grupos)."""
        if isinstance(self.datos, dict):
            for key, value in self.datos.items():
                if isinstance(value, dict):
                    canvas.create_text(50, y, text=str(key), anchor='w', font=('Arial', 10, 'bold'), fill='#1e3a5f')
                    y += 18
                    for k2, v2 in value.items():
                        canvas.create_text(70, y, text=f"  {k2}: {v2}", anchor='w', font=('Arial', 9))
                        y += 16
                    y += 4
                else:
                    canvas.create_text(50, y, text=f"{key}: {value}", anchor='w', font=('Arial', 10))
                    y += 20
        elif isinstance(self.datos, str):
            for line in self.datos.splitlines():
                if line.strip():
                    canvas.create_text(50, y, text=line[:100], anchor='w', font=('Consolas', 9))
                y += 16
    
    def imprimir(self):
        """Función de impresión real"""
        try:
            import win32print
            import win32ui
            from PIL import Image, ImageWin
            
            # Obtener impresora predeterminada
            printer_name = win32print.GetDefaultPrinter()
            
            # Crear contexto de impresora
            hprinter = win32print.OpenPrinter(printer_name)
            try:
                # Crear trabajo de impresión
                hdc = win32ui.CreateDC()
                hdc.CreatePrinterDC(printer_name)
                hdc.StartDoc(self.titulo)
                hdc.StartPage()
                
                # Configurar fuente
                hdc.SetMapMode(win32con.MM_TWIPS)
                
                # Dibujar título
                hdc.TextOut(300, -300, self.titulo)
                
                # Dibujar fecha
                hdc.TextOut(300, -400, f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
                
                hdc.EndPage()
                hdc.EndDoc()
                hdc.DeleteDC()
                
            finally:
                win32print.ClosePrinter(hprinter)
            
            messagebox.showinfo("Impresión", "Documento enviado a la impresora")
            
        except ImportError:
            # Si no hay win32print, guardar como PDF
            self.guardar_como_pdf()
        except Exception as e:
            messagebox.showerror("Error", f"Error al imprimir: {str(e)}")
    
    def guardar_como_pdf(self):
        """Guarda el reporte como PDF con encabezado empresarial."""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.utils import ImageReader
            archivo = tempfile.mktemp(suffix='.pdf', prefix='reporte_')
            c = canvas.Canvas(archivo, pagesize=letter)
            y = 750
            if self.empresa:
                # Logo SSEPI (si existe)
                try:
                    logo_path = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "assets", "ssepi_logo.png"))
                    if os.path.exists(logo_path):
                        img = ImageReader(logo_path)
                        iw, ih = img.getSize()
                        th = 18
                        tw = int((iw / ih) * th) if ih else 70
                        c.drawImage(img, 50, y-18, width=tw, height=th, mask='auto', preserveAspectRatio=True, anchor='c')
                except Exception:
                    pass
                c.setFont("Helvetica-Bold", 12)
                c.drawString(130, y, self.empresa)
                y -= 22
                c.setStrokeColorRGB(0.8, 0.82, 0.88)
                c.line(50, y, 550, y)
                y -= 22
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, y, self.titulo)
            y -= 22
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.39, 0.45, 0.55)
            c.drawString(50, y, f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            y -= 28
            c.setFillColorRGB(0, 0, 0)
            if isinstance(self.datos, dict):
                for key, value in self.datos.items():
                    if isinstance(value, dict):
                        c.setFont("Helvetica-Bold", 10)
                        c.drawString(50, y, str(key))
                        y -= 16
                        c.setFont("Helvetica", 9)
                        for k2, v2 in value.items():
                            c.drawString(70, y, f"  {k2}: {v2}")
                            y -= 14
                        y -= 6
                    else:
                        c.drawString(50, y, f"{key}: {value}")
                        y -= 18
            elif isinstance(self.datos, str):
                c.setFont("Courier", 9)
                for line in self.datos.splitlines():
                    if y < 80:
                        c.showPage()
                        y = 750
                    c.drawString(50, y, line[:90] if len(line) > 90 else line)
                    y -= 14
            c.save()
            os.startfile(archivo)
            messagebox.showinfo("Listo", f"PDF guardado en:\n{archivo}")
        except ImportError:
            messagebox.showerror("Error", "Para guardar PDF instale: pip install reportlab")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar PDF: {str(e)}")
    
    def configurar_pagina(self):
        """Configuración de página"""
        messagebox.showinfo("Configurar página", "Función en desarrollo")

class ImpresionManager:
    def __init__(self):
        pass
    
    def imprimir_balanza(self, parent, datos):
        """Imprime balanza de comprobación"""
        columnas = ["Cuenta", "Tipo / clase", "Descripción", "Saldo Inicial", "Debe", "Haber", "Saldo Final"]
        VistaPreviaImpresion(parent, "Balanza de Comprobación", datos, columnas)
    
    def imprimir_resultados(self, parent, datos):
        """Imprime estado de resultados"""
        VistaPreviaImpresion(parent, "Estado de Resultados", datos)
    
    def imprimir_balance(self, parent, datos):
        """Imprime balance general"""
        VistaPreviaImpresion(parent, "Balance General", datos)
    
    def imprimir_polizas(self, parent, datos):
        """Imprime pólizas"""
        if datos and len(datos[0]) >= 8:
            columnas = ["Fecha", "Tipo", "Número", "Cuentas", "Tipo cuenta", "Concepto", "Total Cargos", "Total Abonos"]
        else:
            columnas = ["Fecha", "Tipo", "Número", "Concepto"]
        VistaPreviaImpresion(parent, "Pólizas del Período", datos, columnas)
    
    def imprimir_catalogo(self, parent, datos):
        """Imprime catálogo de cuentas (incluye Tipo determinación si hay 6 columnas)."""
        # `datos` aquí son los `values` del Treeview (NO incluye el campo #0 de la cuenta).
        # Dependiendo de la cantidad de columnas, ajustamos encabezados para que correspondan.
        if datos:
            ncols = len(datos[0])
        else:
            ncols = 0

        if ncols >= 7:
            columnas = ["Formato COI", "Tipo / clase", "Descripción", "Nivel", "Naturaleza", "Cliente(s)", "Saldo"]
        elif ncols >= 6:
            columnas = ["Tipo / clase", "Descripción", "Nivel", "Naturaleza", "Cliente(s)", "Saldo"]
        elif ncols == 5:
            columnas = ["Tipo / clase", "Descripción", "Nivel", "Naturaleza", "Saldo"]
        else:
            columnas = ["Tipo / clase", "Descripción", "Nivel", "Naturaleza", "Saldo"]
        VistaPreviaImpresion(parent, "Catálogo de Cuentas", datos, columnas)

    def vista_previa_reportes_avanzados(self, parent, titulo, datos, columnas=None, empresa=None):
        """Abre vista previa para Razones financieras o Análisis horizontal (formato empresarial)."""
        VistaPreviaImpresion(parent, titulo, datos, columnas=columnas, empresa=empresa)

    def vista_previa_flujo_efectivo(self, parent, texto_reporte, empresa=None):
        """Abre vista previa del flujo de efectivo (formato empresarial)."""
        VistaPreviaImpresion(parent, "Flujo de efectivo", texto_reporte, columnas=None, empresa=empresa)

    def exportar_auxiliares_coi_pdf(self, parent, datos_auxiliares, meta: dict | None = None):
        """
        Exporta Auxiliares en formato COI-like a PDF (agrupado por cuenta).

        `datos_auxiliares` debe ser la salida de `AuxiliaresManager.reporte_mes()`.
        `meta` puede incluir: empresa, ejercicio, periodo, solo_afectadas.
        """
        meta = meta or {}
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.utils import ImageReader
        except Exception:
            messagebox.showerror("PDF", "Para exportar PDF instale: pip install reportlab")
            return

        empresa = (meta.get("empresa") or "").strip()
        ejercicio = meta.get("ejercicio")
        periodo = meta.get("periodo")
        solo_af = meta.get("solo_afectadas")

        sugerido = f"auxiliares_{periodo:02d}_{ejercicio}.pdf" if isinstance(periodo, int) and isinstance(ejercicio, int) else "auxiliares.pdf"
        ruta = filedialog.asksaveasfilename(
            parent=parent,
            title="Guardar Auxiliares (PDF)",
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF", "*.pdf")],
        )
        if not ruta:
            return

        W, H = landscape(letter)
        c = canvas.Canvas(ruta, pagesize=(W, H))

        def header(titulo: str):
            y = H - 42
            if empresa:
                # Logo SSEPI (si existe)
                try:
                    logo_path = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "assets", "ssepi_logo.png"))
                    if os.path.exists(logo_path):
                        img = ImageReader(logo_path)
                        iw, ih = img.getSize()
                        th = 16
                        tw = int((iw / ih) * th) if ih else 70
                        c.drawImage(img, 40, y - 14, width=tw, height=th, mask="auto", preserveAspectRatio=True)
                except Exception:
                    pass
                c.setFont("Helvetica-Bold", 11)
                c.drawString(40 + 90, y, empresa)
                y -= 18
                c.setStrokeColorRGB(0.8, 0.82, 0.88)
                c.line(40, y, W - 40, y)
                y -= 16

            c.setFont("Helvetica-Bold", 14)
            c.setFillColorRGB(0, 0, 0.5)
            c.drawString(40, y, titulo)
            c.setFillColorRGB(0, 0, 0)
            y -= 18
            c.setFont("Helvetica", 9)
            ley = []
            if isinstance(periodo, int) and isinstance(ejercicio, int):
                ley.append(f"Periodo: {periodo:02d}/{ejercicio}")
            if isinstance(solo_af, bool):
                ley.append("Solo afectadas" if solo_af else "Incluye no afectadas")
            if ley:
                c.setFillColorRGB(0.39, 0.45, 0.55)
                c.drawString(40, y, " · ".join(ley))
                c.setFillColorRGB(0, 0, 0)
            c.setFillColorRGB(0.39, 0.45, 0.55)
            c.drawRightString(W - 40, y, datetime.now().strftime("%d/%m/%Y %H:%M"))
            c.setFillColorRGB(0, 0, 0)
            y -= 18
            c.setStrokeColorRGB(0.85, 0.85, 0.85)
            c.line(40, y, W - 40, y)
            return y - 14

        def nueva_pagina(titulo: str):
            c.showPage()
            return header(titulo)

        y = header("Auxiliares")

        def fmt(x):
            try:
                return f"{float(x or 0):,.2f}"
            except Exception:
                return "0.00"

        for cuenta in (datos_auxiliares or []):
            num = str(cuenta.get("num_cuenta") or "")
            nom = str(cuenta.get("nombre_cuenta") or "")
            si = float(cuenta.get("saldo_inicial") or 0.0)
            tc = float(cuenta.get("total_cargos") or 0.0)
            ta = float(cuenta.get("total_abonos") or 0.0)
            sf = float(cuenta.get("saldo_final") or 0.0)
            movs = cuenta.get("movimientos") or []

            if y < 120:
                y = nueva_pagina("Auxiliares")

            c.setFont("Helvetica-Bold", 11)
            c.setFillColorRGB(0.12, 0.23, 0.37)
            c.drawString(40, y, f"{num}  {nom}".strip())
            c.setFillColorRGB(0, 0, 0)
            y -= 16

            c.setFont("Helvetica", 9)
            c.drawString(40, y, f"Saldo inicial: {fmt(si)}")
            y -= 14

            # Encabezados de tabla
            x_fecha, x_tipo, x_num, x_desc, x_cargo, x_abono, x_saldo = 40, 110, 150, 200, W - 220, W - 140, W - 60
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_fecha, y, "Fecha")
            c.drawString(x_tipo, y, "Tipo")
            c.drawString(x_num, y, "Núm")
            c.drawString(x_desc, y, "Concepto")
            c.drawRightString(x_cargo, y, "Cargo")
            c.drawRightString(x_abono, y, "Abono")
            c.drawRightString(x_saldo, y, "Saldo")
            y -= 10
            c.setStrokeColorRGB(0.7, 0.7, 0.7)
            c.line(40, y, W - 40, y)
            y -= 12

            c.setFont("Helvetica", 9)
            saldo = si
            for m in movs:
                if y < 90:
                    y = nueva_pagina("Auxiliares")
                    # reimprimir encabezado de tabla
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(x_fecha, y, "Fecha")
                    c.drawString(x_tipo, y, "Tipo")
                    c.drawString(x_num, y, "Núm")
                    c.drawString(x_desc, y, "Concepto")
                    c.drawRightString(x_cargo, y, "Cargo")
                    c.drawRightString(x_abono, y, "Abono")
                    c.drawRightString(x_saldo, y, "Saldo")
                    y -= 10
                    c.setStrokeColorRGB(0.7, 0.7, 0.7)
                    c.line(40, y, W - 40, y)
                    y -= 12
                    c.setFont("Helvetica", 9)

                fecha = str(m.get("fecha") or "")[:10]
                tipo = str(m.get("tipo") or "")
                nump = str(m.get("numero") or "")
                desc = str(m.get("descripcion") or "").replace("\n", " ").strip()
                cargo = float(m.get("cargo") or 0.0)
                abono = float(m.get("abono") or 0.0)
                saldo = saldo + cargo - abono

                c.drawString(x_fecha, y, fecha)
                c.drawString(x_tipo, y, tipo[:2])
                c.drawString(x_num, y, nump[:8])
                c.drawString(x_desc, y, desc[:75])
                c.drawRightString(x_cargo, y, fmt(cargo))
                c.drawRightString(x_abono, y, fmt(abono))
                c.drawRightString(x_saldo, y, fmt(saldo))
                y -= 12

            # Totales por cuenta
            y -= 4
            c.setStrokeColorRGB(0.85, 0.85, 0.85)
            c.line(40, y, W - 40, y)
            y -= 14
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_desc, y, "Totales:")
            c.drawRightString(x_cargo, y, fmt(tc))
            c.drawRightString(x_abono, y, fmt(ta))
            c.drawRightString(x_saldo, y, fmt(sf))
            c.setFont("Helvetica", 9)
            y -= 22

        c.save()
        try:
            os.startfile(ruta)
        except Exception:
            pass
        messagebox.showinfo("Auxiliares", f"PDF guardado en:\n{ruta}")

    def exportar_auxiliares_coi_excel(self, parent, datos_auxiliares, meta: dict | None = None):
        """
        Exporta Auxiliares en formato COI-like a Excel (.xlsx), agrupado por cuenta.
        """
        meta = meta or {}
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except Exception:
            messagebox.showerror("Excel", "Para exportar Excel instale: pip install openpyxl")
            return

        empresa = (meta.get("empresa") or "").strip()
        ejercicio = meta.get("ejercicio")
        periodo = meta.get("periodo")
        solo_af = meta.get("solo_afectadas")

        sugerido = f"auxiliares_{periodo:02d}_{ejercicio}.xlsx" if isinstance(periodo, int) and isinstance(ejercicio, int) else "auxiliares.xlsx"
        ruta = filedialog.asksaveasfilename(
            parent=parent,
            title="Guardar Auxiliares (Excel)",
            defaultextension=".xlsx",
            initialfile=sugerido,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not ruta:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Auxiliares"

        bold = Font(bold=True)
        head_fill = PatternFill("solid", fgColor="1E3A5F")
        head_font = Font(bold=True, color="FFFFFF")
        sub_fill = PatternFill("solid", fgColor="E2E8F0")
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        r = 1
        if empresa:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.cell(r, 1, empresa).font = Font(bold=True, size=12)
            r += 1

        titulo = "Auxiliares"
        if isinstance(periodo, int) and isinstance(ejercicio, int):
            titulo += f" {periodo:02d}/{ejercicio}"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(r, 1, titulo).font = Font(bold=True, size=14)
        r += 1

        ley = []
        if isinstance(solo_af, bool):
            ley.append("Solo afectadas" if solo_af else "Incluye no afectadas")
        ley.append(datetime.now().strftime("%d/%m/%Y %H:%M"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(r, 1, " · ".join([x for x in ley if x])).font = Font(size=9, color="64748B")
        r += 2

        headers = ["Fecha", "Tipo", "Núm", "Concepto", "Cargo", "Abono", "Saldo"]
        for cuenta in (datos_auxiliares or []):
            num = str(cuenta.get("num_cuenta") or "")
            nom = str(cuenta.get("nombre_cuenta") or "")
            si = float(cuenta.get("saldo_inicial") or 0.0)
            movs = cuenta.get("movimientos") or []
            tc = float(cuenta.get("total_cargos") or 0.0)
            ta = float(cuenta.get("total_abonos") or 0.0)
            sf = float(cuenta.get("saldo_final") or 0.0)

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            c0 = ws.cell(r, 1, f"{num}  {nom}".strip())
            c0.font = Font(bold=True, color="1E3A5F")
            r += 1

            ws.cell(r, 1, "Saldo inicial").font = bold
            ws.cell(r, 7, si).number_format = "#,##0.00"
            r += 1

            for j, h in enumerate(headers, start=1):
                cell = ws.cell(r, j, h)
                cell.font = head_font
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            r += 1

            saldo = si
            for m in movs:
                fecha = str(m.get("fecha") or "")[:10]
                tipo = str(m.get("tipo") or "")
                nump = str(m.get("numero") or "")
                desc = str(m.get("descripcion") or "").replace("\n", " ").strip()
                cargo = float(m.get("cargo") or 0.0)
                abono = float(m.get("abono") or 0.0)
                saldo = saldo + cargo - abono

                ws.cell(r, 1, fecha)
                ws.cell(r, 2, tipo)
                ws.cell(r, 3, nump)
                ws.cell(r, 4, desc)
                ws.cell(r, 5, cargo).number_format = "#,##0.00"
                ws.cell(r, 6, abono).number_format = "#,##0.00"
                ws.cell(r, 7, saldo).number_format = "#,##0.00"
                for j in range(1, 8):
                    ws.cell(r, j).border = border
                r += 1

            # Totales
            for j in range(1, 8):
                ws.cell(r, j).fill = sub_fill
                ws.cell(r, j).border = border
                ws.cell(r, j).font = bold
            ws.cell(r, 4, "Totales:")
            ws.cell(r, 5, tc).number_format = "#,##0.00"
            ws.cell(r, 6, ta).number_format = "#,##0.00"
            ws.cell(r, 7, sf).number_format = "#,##0.00"
            r += 2

        # Anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14

        try:
            wb.save(ruta)
        except Exception as e:
            messagebox.showerror("Excel", f"No se pudo guardar el archivo:\n{e}")
            return

        try:
            os.startfile(ruta)
        except Exception:
            pass
        messagebox.showinfo("Auxiliares", f"Excel guardado en:\n{ruta}")