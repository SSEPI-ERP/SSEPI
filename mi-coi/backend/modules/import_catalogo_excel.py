from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
import csv

from backend.models.catalogo import CatalogoCuentas


def _norm(s: Any) -> str:
    return str(s or "").strip()


def _norm_key(s: Any) -> str:
    return _norm(s).lower().replace(" ", "").replace("_", "").replace("-", "")


def _to_int(val: Any, default: int = 0) -> int:
    try:
        return int(float(str(val).strip()))
    except Exception:
        return default


def _to_float(val: Any, default: float = 0.0) -> float:
    try:
        s = str(val).strip().replace(",", "")
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def _bool_from_any(val: Any, default: bool = True) -> bool:
    if val is None:
        return default
    s = str(val).strip().lower()
    if s in ("1", "true", "t", "si", "sí", "s", "y", "yes", "activa", "activo"):
        return True
    if s in ("0", "false", "f", "no", "n", "inactiva", "inactivo"):
        return False
    return default


@dataclass
class ImportStats:
    leidas: int = 0
    ok: int = 0
    creadas: int = 0
    actualizadas: int = 0
    omitidas: int = 0
    errores: int = 0


class ImportCatalogoExcel:
    """
    Importa catálogo desde Excel (.xlsx).

    Encabezados soportados (flexibles):
    - num_cuenta / cuenta / numero / código
    - nombre_cuenta / nombre / descripcion
    - nivel
    - naturaleza (DEUDORA/ACREEDORA)
    - cuenta_mayor / mayor / padre
    - tipo_cuenta (DETALLE/ACUMULATIVA)
    - moneda
    - codigo_agrupador_sat / sat / codigo_sat
    - no_incluir_xml (0/1)
    - rubro_financiero
    - rubro_diot
    - activa (0/1)
    - limite_credito_mxn
    """

    def __init__(self, catalogo: Optional[CatalogoCuentas] = None):
        self.catalogo = catalogo or CatalogoCuentas()

    def _cargar_filas_archivo(self, path: str) -> Any:
        """Devuelve (headers_row, data_rows) o dict error."""
        if not os.path.isfile(path):
            return {"exito": False, "error": "Archivo no encontrado"}
        low = path.lower()
        if low.endswith(".csv"):
            try:
                with open(path, newline="", encoding="utf-8-sig") as f:
                    rd = csv.reader(f)
                    rows = list(rd)
            except Exception as e:
                return {"exito": False, "error": str(e)}
            if not rows:
                return {"exito": False, "error": "El archivo está vacío"}
            return rows[0], rows[1:]
        try:
            from openpyxl import load_workbook
        except Exception as e:
            return {"exito": False, "error": f"Falta openpyxl: {e}"}
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"exito": False, "error": "El archivo está vacío"}
        return rows[0], rows[1:]

    def leer_preview(self, file_path: str, *, max_rows: int = 50) -> Dict[str, Any]:
        loaded = self._cargar_filas_archivo(file_path)
        if isinstance(loaded, dict):
            return loaded
        hdr_row, data_rows = loaded
        headers = [str(c or "").strip() for c in hdr_row]
        data = []
        for r in data_rows[: int(max_rows)]:
            data.append([("" if v is None else v) for v in r[: len(headers)]])
        return {"exito": True, "headers": headers, "rows": data, "total_rows": max(0, len(data_rows))}

    def _ensure_padre_recursivo(
        self,
        cuenta_mayor: Optional[str],
        *,
        crear: bool,
        errores: List[str],
        row_tag: str,
    ) -> bool:
        padre = _norm(cuenta_mayor) or None
        if not padre:
            return True
        if self.catalogo.obtener_cuenta(padre):
            return True
        if not crear:
            errores.append(f"{row_tag}: Falta cuenta padre «{padre}».")
            return False
        if "." not in padre:
            errores.append(f"{row_tag}: No se puede crear padre automático sin segmentos (.) en «{padre}».")
            return False
        partes = [p for p in padre.split(".") if p.strip()]
        abuelo = ".".join(partes[:-1]) if len(partes) > 1 else None
        if abuelo and not self._ensure_padre_recursivo(abuelo, crear=crear, errores=errores, row_tag=row_tag):
            return False
        if self.catalogo.obtener_cuenta(padre):
            return True
        nivel_p = max(1, len(partes))
        stub = {
            "num_cuenta": padre,
            "nombre_cuenta": f"(Import) {padre}",
            "nivel": nivel_p,
            "naturaleza": "DEUDORA",
            "cuenta_mayor": abuelo,
            "tipo_cuenta": "ACUMULATIVA",
            "moneda": "MXN",
            "codigo_agrupador_sat": None,
            "no_incluir_xml": False,
            "rubro_financiero": None,
            "rubro_diot": None,
            "activa": True,
            "saldo_inicial": 0,
            "saldo_final": 0,
            "limite_credito_mxn": None,
        }
        ok, _msg = self.catalogo.agregar_cuenta_completa(stub)
        return bool(ok)

    def importar(
        self,
        file_path: str,
        *,
        modo: str = "upsert",  # upsert|insert_only|update_only
        permitir_incompletas: bool = False,
        crear_padres_faltantes: bool = False,
    ) -> Dict[str, Any]:
        loaded = self._cargar_filas_archivo(file_path)
        if isinstance(loaded, dict):
            return loaded
        hdr, body = loaded
        rows = [hdr] + list(body)
        if len(rows) < 2:
            return {"exito": False, "error": "No hay filas de datos (solo encabezados)."}
        hdr = [c for c in rows[0]]
        idx: Dict[str, int] = {}
        for i, h in enumerate(hdr):
            k = _norm_key(h)
            if not k:
                continue
            idx[k] = i

        def pick(*keys: str) -> Optional[int]:
            for k in keys:
                k2 = _norm_key(k)
                if k2 in idx:
                    return idx[k2]
            return None

        col_num = pick("num_cuenta", "cuenta", "numero", "codigo", "código", "no_cuenta")
        col_nom = pick("nombre_cuenta", "nombre", "descripcion", "descripción")
        if col_num is None or col_nom is None:
            return {
                "exito": False,
                "error": "Encabezados requeridos: num_cuenta/cuenta y nombre_cuenta/nombre.",
                "headers": [str(x or "") for x in hdr],
            }

        col_nivel = pick("nivel")
        col_nat = pick("naturaleza", "nat")
        col_mayor = pick("cuenta_mayor", "mayor", "padre", "cuentapadre")
        col_tipo = pick("tipo_cuenta", "tipo", "detalle")
        col_mon = pick("moneda", "iso")
        col_sat = pick("codigo_agrupador_sat", "codigo_sat", "sat", "agrupadorsat", "codigosat")
        col_no_xml = pick("no_incluir_xml", "noincluirxml", "incluirxml")
        col_rfin = pick("rubro_financiero", "rubrofinanciero")
        col_rdiot = pick("rubro_diot", "rubrodiot", "tipo_operacion_diot", "tipooperaciondiot")
        col_act = pick("activa", "activo", "estatus")
        col_lim = pick("limite_credito_mxn", "limite", "credito", "limitecredito")
        col_tb = pick("tipo_balance", "tipobalance", "clase_balance")

        def _depth_codigo(nm: str) -> int:
            s = _norm(nm)
            if not s:
                return 999
            if "." in s:
                return len([p for p in s.split(".") if p.strip()])
            return 1

        indexed: List[Tuple[int, List[Any]]] = [(i + 2, list(row)) for i, row in enumerate(rows[1:])]
        indexed.sort(
            key=lambda it: (
                _depth_codigo(it[1][col_num] if col_num is not None and col_num < len(it[1]) else ""),
                _norm(it[1][col_num] if col_num is not None and col_num < len(it[1]) else ""),
            ),
        )

        stats = ImportStats()
        errores: List[str] = []

        for row_i, r in indexed:
            stats.leidas += 1
            num = _norm(r[col_num] if col_num is not None and col_num < len(r) else "")
            nom = _norm(r[col_nom] if col_nom is not None and col_nom < len(r) else "")
            if not num or not nom:
                stats.omitidas += 1
                continue

            nivel = _to_int(r[col_nivel], 0) if (col_nivel is not None and col_nivel < len(r)) else 0
            if nivel <= 0:
                # Inferir nivel básico por separadores COI (4-3-3) o por longitud
                if "." in num:
                    nivel = len([p for p in num.split(".") if p.strip()])
                else:
                    n2 = num.replace("-", "").strip()
                    if len(n2) <= 4:
                        nivel = 1
                    elif len(n2) <= 7:
                        nivel = 2
                    elif len(n2) <= 10:
                        nivel = 3
                    else:
                        nivel = 4

            naturaleza = _norm(r[col_nat] if (col_nat is not None and col_nat < len(r)) else "").upper()
            if naturaleza not in ("DEUDORA", "ACREEDORA"):
                naturaleza = "DEUDORA"

            cuenta_mayor = _norm(r[col_mayor] if (col_mayor is not None and col_mayor < len(r)) else "") or None
            if not cuenta_mayor and "." in num:
                cuenta_mayor = num.rsplit(".", 1)[0].strip() or None
            tipo_cuenta = _norm(r[col_tipo] if (col_tipo is not None and col_tipo < len(r)) else "").upper()
            if tipo_cuenta not in ("DETALLE", "ACUMULATIVA"):
                tipo_cuenta = "DETALLE" if nivel >= 3 else "ACUMULATIVA"

            moneda = _norm(r[col_mon] if (col_mon is not None and col_mon < len(r)) else "") or "MXN"
            codigo_sat = _norm(r[col_sat] if (col_sat is not None and col_sat < len(r)) else "") or None

            no_incluir_xml = False
            if col_no_xml is not None and col_no_xml < len(r):
                v = r[col_no_xml]
                # Si traen "incluirxml": invertimos si parece booleano "si/no"
                k = _norm_key(hdr[col_no_xml])
                if k in ("incluirxml",):
                    no_incluir_xml = not _bool_from_any(v, True)
                else:
                    no_incluir_xml = _bool_from_any(v, False)

            rubro_financiero = _norm(r[col_rfin] if (col_rfin is not None and col_rfin < len(r)) else "") or None
            if col_tb is not None and col_tb < len(r):
                tbv = _norm(r[col_tb])
                if tbv:
                    rubro_financiero = tbv
            rubro_diot = _norm(r[col_rdiot] if (col_rdiot is not None and col_rdiot < len(r)) else "") or None
            activa = _bool_from_any(r[col_act], True) if (col_act is not None and col_act < len(r)) else True
            limite_credito_mxn = _to_float(r[col_lim], 0.0) if (col_lim is not None and col_lim < len(r)) else None

            datos = {
                "num_cuenta": num,
                "nombre_cuenta": nom,
                "nivel": nivel,
                "naturaleza": naturaleza,
                "cuenta_mayor": cuenta_mayor,
                "tipo_cuenta": tipo_cuenta,
                "moneda": moneda,
                "codigo_agrupador_sat": codigo_sat,
                "no_incluir_xml": bool(no_incluir_xml),
                "rubro_financiero": rubro_financiero,
                "rubro_diot": rubro_diot,
                "activa": bool(activa),
                "saldo_inicial": 0,
                "saldo_final": 0,
                "limite_credito_mxn": limite_credito_mxn,
            }

            # Validación mínima: si es DETALLE, normalmente SAT requerido
            if not permitir_incompletas and datos["tipo_cuenta"] == "DETALLE" and not datos.get("codigo_agrupador_sat"):
                stats.errores += 1
                errores.append(f"Fila {row_i}: Cuenta {num} DETALLE sin código SAT.")
                continue

            existente = self.catalogo.obtener_cuenta(num)
            if existente:
                if modo == "insert_only":
                    stats.omitidas += 1
                    continue
                ok, msg = self.catalogo.actualizar_cuenta_completa(num, datos)
                if ok:
                    stats.ok += 1
                    stats.actualizadas += 1
                else:
                    stats.errores += 1
                    errores.append(f"Fila {row_i}: {num} {msg}")
            else:
                if modo == "update_only":
                    stats.omitidas += 1
                    continue
                if datos.get("cuenta_mayor") and not self._ensure_padre_recursivo(
                    datos.get("cuenta_mayor"),
                    crear=bool(crear_padres_faltantes),
                    errores=errores,
                    row_tag=f"Fila {row_i}",
                ):
                    stats.errores += 1
                    continue
                ok, msg = self.catalogo.agregar_cuenta_completa(datos)
                if ok:
                    stats.ok += 1
                    stats.creadas += 1
                else:
                    stats.errores += 1
                    errores.append(f"Fila {row_i}: {num} {msg}")

        return {"exito": True, "stats": stats.__dict__, "errores": errores[:200]}

