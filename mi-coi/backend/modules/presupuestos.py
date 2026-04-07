"""
Presupuestos anuales por cuenta (12 meses) y comparativo vs real.

- Captura: monto anual repartido en 12 meses (uniforme, por % o manual).
- Comparativo: mes y acumulado (YTD), variación y % cumplimiento.
- Importación / exportación Excel (12 columnas mensuales).
"""

from __future__ import annotations

import sqlite3
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from config import get_db_path


class PresupuestosManager:
    def __init__(self, db_path: str | None = None):
        self.db_path = db_path or get_db_path()
        self._ensure_schema()

    def _ensure_schema(self) -> None:
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS presupuestos_encabezado (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ejercicio INTEGER NOT NULL,
                    periodo INTEGER NOT NULL,
                    usuario TEXT NOT NULL,
                    fecha_creacion TEXT NOT NULL,
                    UNIQUE(ejercicio, periodo)
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS presupuestos_detalle (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    id_presupuesto INTEGER NOT NULL,
                    num_cuenta TEXT NOT NULL,
                    monto_presupuestado_mn REAL NOT NULL DEFAULT 0,
                    moneda TEXT NOT NULL DEFAULT 'MXN',
                    fecha_actualizacion TEXT NOT NULL,
                    UNIQUE(id_presupuesto, num_cuenta),
                    FOREIGN KEY(id_presupuesto) REFERENCES presupuestos_encabezado(id) ON DELETE CASCADE
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS presupuesto_anual_cuenta (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ejercicio INTEGER NOT NULL,
                    num_cuenta TEXT NOT NULL,
                    m01 REAL NOT NULL DEFAULT 0,
                    m02 REAL NOT NULL DEFAULT 0,
                    m03 REAL NOT NULL DEFAULT 0,
                    m04 REAL NOT NULL DEFAULT 0,
                    m05 REAL NOT NULL DEFAULT 0,
                    m06 REAL NOT NULL DEFAULT 0,
                    m07 REAL NOT NULL DEFAULT 0,
                    m08 REAL NOT NULL DEFAULT 0,
                    m09 REAL NOT NULL DEFAULT 0,
                    m10 REAL NOT NULL DEFAULT 0,
                    m11 REAL NOT NULL DEFAULT 0,
                    m12 REAL NOT NULL DEFAULT 0,
                    usuario TEXT,
                    actualizado_en TEXT NOT NULL,
                    UNIQUE(ejercicio, num_cuenta)
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS presupuesto_migracion_legacy (
                    hecho INTEGER NOT NULL DEFAULT 1
                )
                """
            )
            conn.commit()
        self._migrate_legacy_presupuestos()

    def _migrate_legacy_presupuestos(self) -> None:
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1 FROM presupuesto_migracion_legacy LIMIT 1")
                if cur.fetchone():
                    return
                cur.execute(
                    """
                    SELECT e.ejercicio, d.num_cuenta, e.periodo, d.monto_presupuestado_mn
                    FROM presupuestos_detalle d
                    JOIN presupuestos_encabezado e ON e.id = d.id_presupuesto
                    WHERE e.periodo BETWEEN 1 AND 12
                    """
                )
                rows = cur.fetchall()
                agg: Dict[Tuple[int, str], List[float]] = defaultdict(lambda: [0.0] * 12)
                for ej, nc, per, monto in rows:
                    agg[(int(ej), str(nc).strip())][int(per) - 1] = float(monto or 0)
                ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                for (ej, nc), meses in agg.items():
                    placeholders = ",".join(["?"] * 12)
                    cur.execute(
                        f"""
                        INSERT INTO presupuesto_anual_cuenta (
                            ejercicio, num_cuenta, m01,m02,m03,m04,m05,m06,m07,m08,m09,m10,m11,m12,
                            usuario, actualizado_en
                        ) VALUES (?, ?, {placeholders}, ?, ?)
                        ON CONFLICT(ejercicio, num_cuenta) DO UPDATE SET
                            m01 = presupuesto_anual_cuenta.m01 + excluded.m01,
                            m02 = presupuesto_anual_cuenta.m02 + excluded.m02,
                            m03 = presupuesto_anual_cuenta.m03 + excluded.m03,
                            m04 = presupuesto_anual_cuenta.m04 + excluded.m04,
                            m05 = presupuesto_anual_cuenta.m05 + excluded.m05,
                            m06 = presupuesto_anual_cuenta.m06 + excluded.m06,
                            m07 = presupuesto_anual_cuenta.m07 + excluded.m07,
                            m08 = presupuesto_anual_cuenta.m08 + excluded.m08,
                            m09 = presupuesto_anual_cuenta.m09 + excluded.m09,
                            m10 = presupuesto_anual_cuenta.m10 + excluded.m10,
                            m11 = presupuesto_anual_cuenta.m11 + excluded.m11,
                            m12 = presupuesto_anual_cuenta.m12 + excluded.m12,
                            actualizado_en = excluded.actualizado_en
                        """,
                        (ej, nc, *meses, "Migración legacy", ahora),
                    )
                cur.execute("INSERT INTO presupuesto_migracion_legacy (hecho) VALUES (1)")
                conn.commit()
        except Exception:
            pass

    @staticmethod
    def _sum_meses(row: Sequence[float]) -> float:
        return float(sum(float(x or 0) for x in row))

    def listar_cuentas_presupuesto_anual(self, ejercicio: int) -> List[Dict[str, Any]]:
        ejercicio = int(ejercicio)
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                """
                SELECT p.*, COALESCE(c.nombre_cuenta,'') AS nombre_cuenta
                FROM presupuesto_anual_cuenta p
                LEFT JOIN catalogo_cuentas c ON c.num_cuenta = p.num_cuenta
                WHERE p.ejercicio = ?
                ORDER BY p.num_cuenta
                """,
                (ejercicio,),
            )
            out = []
            for r in cur.fetchall():
                d = dict(r)
                meses = [float(d.get(f"m{m:02d}") or 0) for m in range(1, 13)]
                d["meses"] = meses
                d["monto_anual"] = self._sum_meses(meses)
                out.append(d)
            return out

    def obtener_linea_anual(self, ejercicio: int, num_cuenta: str) -> Optional[Dict[str, Any]]:
        num_cuenta = str(num_cuenta or "").strip()
        if not num_cuenta:
            return None
        rows = [x for x in self.listar_cuentas_presupuesto_anual(ejercicio) if x.get("num_cuenta") == num_cuenta]
        return rows[0] if rows else None

    def guardar_meses_cuenta(
        self,
        ejercicio: int,
        num_cuenta: str,
        meses: Sequence[float],
        usuario: str,
    ) -> Dict[str, Any]:
        num_cuenta = str(num_cuenta or "").strip()
        if not num_cuenta:
            return {"exito": False, "error": "num_cuenta requerido"}
        if len(meses) != 12:
            return {"exito": False, "error": "Se requieren 12 montos mensuales"}
        vals = [float(meses[i] or 0) for i in range(12)]
        ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute(
                f"""
                INSERT INTO presupuesto_anual_cuenta (
                    ejercicio, num_cuenta,
                    m01,m02,m03,m04,m05,m06,m07,m08,m09,m10,m11,m12,
                    usuario, actualizado_en
                ) VALUES (?, ?, ?,?,?,?,?,?,?,?,?,?,?, ?, ?)
                ON CONFLICT(ejercicio, num_cuenta) DO UPDATE SET
                    m01=excluded.m01, m02=excluded.m02, m03=excluded.m03, m04=excluded.m04,
                    m05=excluded.m05, m06=excluded.m06, m07=excluded.m07, m08=excluded.m08,
                    m09=excluded.m09, m10=excluded.m10, m11=excluded.m11, m12=excluded.m12,
                    usuario=excluded.usuario, actualizado_en=excluded.actualizado_en
                """,
                (int(ejercicio), num_cuenta, *vals, usuario or "Sistema", ahora),
            )
            conn.commit()
        return {"exito": True}

    def distribuir_uniforme(self, ejercicio: int, num_cuenta: str, monto_anual: float, usuario: str) -> Dict[str, Any]:
        monto_anual = float(monto_anual or 0)
        base = round(monto_anual / 12.0, 2)
        meses = [base] * 12
        diff = round(monto_anual - sum(meses), 2)
        meses[11] = round(meses[11] + diff, 2)
        return self.guardar_meses_cuenta(ejercicio, num_cuenta, meses, usuario)

    def distribuir_porcentajes(
        self,
        ejercicio: int,
        num_cuenta: str,
        monto_anual: float,
        porcentajes: Sequence[float],
        usuario: str,
    ) -> Dict[str, Any]:
        monto_anual = float(monto_anual or 0)
        if len(porcentajes) != 12:
            return {"exito": False, "error": "Se requieren 12 porcentajes"}
        p = [max(0.0, float(x or 0)) for x in porcentajes]
        s = sum(p)
        if s <= 0:
            return {"exito": False, "error": "La suma de porcentajes debe ser mayor que cero"}
        p = [x / s for x in p]
        meses = [round(monto_anual * p[i], 2) for i in range(12)]
        diff = round(monto_anual - sum(meses), 2)
        meses[11] = round(meses[11] + diff, 2)
        return self.guardar_meses_cuenta(ejercicio, num_cuenta, meses, usuario)

    def eliminar_linea_anual(self, ejercicio: int, num_cuenta: str) -> Dict[str, Any]:
        num_cuenta = str(num_cuenta or "").strip()
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                "DELETE FROM presupuesto_anual_cuenta WHERE ejercicio = ? AND num_cuenta = ?",
                (int(ejercicio), num_cuenta),
            )
            conn.commit()
        return {"exito": True}

    def _naturaleza_cuenta(self, cur: sqlite3.Cursor, num_cuenta: str) -> str:
        cur.execute(
            "SELECT UPPER(COALESCE(naturaleza,'DEUDORA')) FROM catalogo_cuentas WHERE num_cuenta = ? LIMIT 1",
            (str(num_cuenta).strip(),),
        )
        row = cur.fetchone()
        return str(row[0] or "DEUDORA").strip() if row else "DEUDORA"

    def _neto_real_cuenta_mes(self, cur: sqlite3.Cursor, ejercicio: int, mes: int, num_cuenta: str) -> float:
        """Movimiento del mes en sentido 'natural' (deudora: cargo-abono; acreedora: abono-cargo), solo pólizas afectadas."""
        nc = str(num_cuenta).strip()
        cur.execute(
            """
            SELECT COALESCE(SUM(COALESCE(pp.cargo_mn, pp.cargo, 0)), 0),
                   COALESCE(SUM(COALESCE(pp.abono_mn, pp.abono, 0)), 0)
            FROM partidas_poliza pp
            JOIN polizas p ON p.id = pp.id_poliza
            WHERE pp.num_cuenta = ?
              AND CAST(strftime('%Y', p.fecha) AS INTEGER) = ?
              AND CAST(strftime('%m', p.fecha) AS INTEGER) = ?
              AND UPPER(COALESCE(p.estatus,'A')) = 'A'
            """,
            (nc, int(ejercicio), int(mes)),
        )
        row = cur.fetchone()
        c_part, a_part = float(row[0] or 0), float(row[1] or 0)
        cur.execute(
            """
            SELECT COALESCE(SUM(COALESCE(m.cargo, 0)), 0),
                   COALESCE(SUM(COALESCE(m.abono, 0)), 0)
            FROM movimientos m
            JOIN polizas p ON p.id = m.poliza_id
            WHERE m.num_cuenta = ?
              AND CAST(strftime('%Y', p.fecha) AS INTEGER) = ?
              AND CAST(strftime('%m', p.fecha) AS INTEGER) = ?
              AND UPPER(COALESCE(p.estatus,'A')) = 'A'
              AND NOT EXISTS (SELECT 1 FROM partidas_poliza px WHERE px.id_poliza = p.id)
            """,
            (nc, int(ejercicio), int(mes)),
        )
        row2 = cur.fetchone()
        c = c_part + float(row2[0] or 0)
        a = a_part + float(row2[1] or 0)
        nat = self._naturaleza_cuenta(cur, nc)
        if nat == "ACREEDORA":
            return float(a - c)
        return float(c - a)

    def comparativo_extendido(self, ejercicio: int, mes: int) -> List[Dict[str, Any]]:
        """
        Filas: presupuesto mes, real mes, var mes; pres YTD, real YTD, var YTD, % cumplimiento YTD.
        """
        ejercicio = int(ejercicio)
        mes = int(mes)
        if mes < 1 or mes > 12:
            return []
        lineas = self.listar_cuentas_presupuesto_anual(ejercicio)
        if not lineas:
            return []
        out: List[Dict[str, Any]] = []
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            for ln in lineas:
                nc = str(ln.get("num_cuenta") or "").strip()
                meses = ln.get("meses") or [0.0] * 12
                pres_mes = float(meses[mes - 1] or 0)
                pres_ytd = float(sum(float(meses[i] or 0) for i in range(mes)))
                real_mes = self._neto_real_cuenta_mes(cur, ejercicio, mes, nc)
                real_ytd = sum(self._neto_real_cuenta_mes(cur, ejercicio, m, nc) for m in range(1, mes + 1))
                var_mes = real_mes - pres_mes
                var_ytd = real_ytd - pres_ytd
                if abs(pres_ytd) >= 0.005:
                    pct = (real_ytd / pres_ytd) * 100.0
                else:
                    pct = None
                out.append(
                    {
                        "num_cuenta": nc,
                        "nombre_cuenta": ln.get("nombre_cuenta", ""),
                        "presupuesto_mes": pres_mes,
                        "real_mes": real_mes,
                        "variacion_mes": var_mes,
                        "presupuesto_ytd": pres_ytd,
                        "real_ytd": real_ytd,
                        "variacion_ytd": var_ytd,
                        "pct_cumplimiento_ytd": pct,
                    }
                )
        return out

    # --- Compatibilidad API anterior (un mes) ---
    def _get_or_create_encabezado(self, ejercicio: int, periodo: int, usuario: str) -> int:
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO presupuestos_encabezado (ejercicio, periodo, usuario, fecha_creacion)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(ejercicio, periodo) DO UPDATE SET
                    usuario=excluded.usuario,
                    fecha_creacion=excluded.fecha_creacion
                """,
                (ejercicio, periodo, usuario or "Sistema", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            cur.execute(
                "SELECT id FROM presupuestos_encabezado WHERE ejercicio=? AND periodo=?",
                (ejercicio, periodo),
            )
            row = cur.fetchone()
            return int(row[0]) if row else 0

    def establecer_monto(self, ejercicio: int, periodo: int, num_cuenta: str, monto: float, usuario: str) -> Dict[str, Any]:
        """Actualiza solo el mes indicado en la tabla anual (compatibilidad)."""
        num_cuenta = str(num_cuenta or "").strip()
        if not num_cuenta:
            return {"exito": False, "error": "num_cuenta requerido"}
        periodo = int(periodo)
        if periodo < 1 or periodo > 12:
            return {"exito": False, "error": "periodo debe ser 1-12"}
        ex = self.obtener_linea_anual(ejercicio, num_cuenta)
        meses = list(ex.get("meses") if ex else [0.0] * 12)
        meses[periodo - 1] = float(monto or 0)
        r = self.guardar_meses_cuenta(ejercicio, num_cuenta, meses, usuario)
        if r.get("exito"):
            try:
                pid = self._get_or_create_encabezado(int(ejercicio), periodo, usuario)
                with sqlite3.connect(self.db_path) as conn:
                    cur = conn.cursor()
                    cur.execute(
                        """
                        INSERT INTO presupuestos_detalle (id_presupuesto, num_cuenta, monto_presupuestado_mn, moneda, fecha_actualizacion)
                        VALUES (?, ?, ?, 'MXN', ?)
                        ON CONFLICT(id_presupuesto, num_cuenta) DO UPDATE SET
                            monto_presupuestado_mn=excluded.monto_presupuestado_mn,
                            fecha_actualizacion=excluded.fecha_actualizacion
                        """,
                        (pid, num_cuenta, float(monto or 0), datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                    )
                    conn.commit()
            except Exception:
                pass
        return r

    def obtener_detalle(self, ejercicio: int, periodo: int) -> List[Dict[str, Any]]:
        periodo = int(periodo)
        rows = self.listar_cuentas_presupuesto_anual(ejercicio)
        out = []
        for ln in rows:
            meses = ln.get("meses") or [0] * 12
            m = float(meses[periodo - 1] if 1 <= periodo <= 12 else 0)
            out.append(
                {
                    "num_cuenta": ln.get("num_cuenta"),
                    "monto_presupuestado_mn": m,
                    "nombre_cuenta": ln.get("nombre_cuenta", ""),
                }
            )
        return [x for x in out if abs(float(x.get("monto_presupuestado_mn") or 0)) >= 0.005]

    def _real_por_cuenta_mn(self, ejercicio: int, periodo: int, num_cuenta: str) -> float:
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            return self._neto_real_cuenta_mes(cur, int(ejercicio), int(periodo), num_cuenta)

    def comparativo(self, ejercicio: int, periodo: int) -> List[Dict[str, Any]]:
        detalle = self.obtener_detalle(ejercicio, periodo)
        if not detalle:
            return []
        out: List[Dict[str, Any]] = []
        for d in detalle:
            nc = d.get("num_cuenta") or ""
            presup = float(d.get("monto_presupuestado_mn") or 0)
            real = self._real_por_cuenta_mn(int(ejercicio), int(periodo), nc)
            out.append(
                {
                    "num_cuenta": nc,
                    "nombre_cuenta": d.get("nombre_cuenta", ""),
                    "presupuesto": presup,
                    "real": float(real),
                    "diferencia": float(real - presup),
                }
            )
        return out

    def exportar_excel(self, ejercicio: int, ruta: str) -> Dict[str, Any]:
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"exito": False, "error": "openpyxl no instalado"}
        ejercicio = int(ejercicio)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Pto{ejercicio}"
        hdr = ["num_cuenta", "nombre_cuenta"] + [f"mes_{i:02d}" for i in range(1, 13)]
        ws.append(hdr)
        for ln in self.listar_cuentas_presupuesto_anual(ejercicio):
            meses = ln.get("meses") or [0] * 12
            ws.append([ln.get("num_cuenta"), ln.get("nombre_cuenta", "")] + meses)
        wb.save(ruta)
        return {"exito": True, "archivo": ruta}

    def importar_excel(self, ejercicio: int, ruta: str, usuario: str) -> Dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"exito": False, "error": "openpyxl no instalado"}
        ejercicio = int(ejercicio)
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return {"exito": False, "error": str(e)}
        if not rows:
            return {"exito": False, "error": "Hoja vacía"}
        header = [str(c or "").strip().lower() for c in rows[0]]
        def idx(name: str) -> int:
            for i, h in enumerate(header):
                if h.replace(" ", "_") == name or h == name:
                    return i
            return -1

        i_cuenta = idx("num_cuenta")
        if i_cuenta < 0:
            i_cuenta = 0
        mes_idx = []
        for i in range(1, 13):
            j = idx(f"mes_{i:02d}")
            if j < 0:
                j = idx(f"mes{i}")
            if j < 0 and len(header) >= i_cuenta + 1 + i:
                j = i_cuenta + i
            mes_idx.append(j)
        cargados = 0
        errores: List[str] = []
        for ridx, row in enumerate(rows[1:], start=2):
            if not row or row[i_cuenta] is None:
                continue
            nc = str(row[i_cuenta]).strip()
            if not nc or nc.lower() == "num_cuenta":
                continue
            vals = []
            for j in mes_idx:
                if j < 0 or j >= len(row):
                    vals.append(0.0)
                else:
                    try:
                        vals.append(float(row[j] or 0))
                    except (TypeError, ValueError):
                        vals.append(0.0)
            rr = self.guardar_meses_cuenta(ejercicio, nc, vals, usuario)
            if rr.get("exito"):
                cargados += 1
            else:
                errores.append(f"Fila {ridx} {nc}: {rr.get('error')}")
        return {"exito": True, "cargados": cargados, "errores": errores}
