# backend/models/polizas.py
import json
import sqlite3
from collections import Counter
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
try:
    from config import get_db_path
except ImportError:
    def get_db_path():
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        return os.path.join(base_dir, 'backend', 'database', 'contabilidad.db')
try:
    from backend.modules.periodos_bloqueados import PeriodosBloqueados
except Exception:
    try:
        from modules.periodos_bloqueados import PeriodosBloqueados
    except Exception:
        PeriodosBloqueados = None

class SistemaPolizas:
    def __init__(self, db_path: str = None):
        self.db_path = db_path if db_path else get_db_path()
        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        self._ensure_schema()

    def _ensure_schema(self):
        """Crea tablas mínimas si no existen (polizas, movimientos, saldos_mensuales)."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("PRAGMA foreign_keys = ON;")
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS polizas (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        numero_poliza INTEGER NOT NULL,
                        tipo_poliza TEXT NOT NULL,
                        fecha TEXT NOT NULL,
                        concepto TEXT NOT NULL,
                        moneda TEXT DEFAULT 'MXN',
                        tipo_cambio REAL DEFAULT 1.0,
                        estatus TEXT DEFAULT 'C'
                    );
                """)
                cur.execute("PRAGMA table_info(polizas)")
                pcols = [r[1] for r in cur.fetchall()]
                for col, ddl in (
                    ("moneda", "TEXT DEFAULT 'MXN'"),
                    ("tipo_cambio", "REAL DEFAULT 1.0"),
                    ("estatus", "TEXT DEFAULT 'C'"),
                ):
                    if col not in pcols:
                        try:
                            cur.execute(f"ALTER TABLE polizas ADD COLUMN {col} {ddl}")
                        except sqlite3.OperationalError:
                            pass
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS movimientos (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        poliza_id INTEGER NOT NULL,
                        num_cuenta TEXT NOT NULL,
                        concepto_mov TEXT,
                        cargo REAL DEFAULT 0,
                        abono REAL DEFAULT 0,
                        cliente_rfc TEXT,
                        cliente_nombre TEXT,
                        FOREIGN KEY(poliza_id) REFERENCES polizas(id) ON DELETE CASCADE
                    );
                """)
                # Migración: añadir columnas cliente si la tabla ya existía sin ellas
                cur.execute("PRAGMA table_info(movimientos)")
                cols = [r[1] for r in cur.fetchall()]
                for col in ("cliente_rfc", "cliente_nombre"):
                    if col not in cols:
                        try:
                            cur.execute(f"ALTER TABLE movimientos ADD COLUMN {col} TEXT")
                        except sqlite3.OperationalError:
                            pass
                cur.execute("PRAGMA table_info(movimientos)")
                cols = [r[1] for r in cur.fetchall()]
                for col, ddl in (
                    ("centro_costo_id", "INTEGER"),
                    ("numero_linea", "INTEGER"),
                ):
                    if col not in cols:
                        try:
                            cur.execute(f"ALTER TABLE movimientos ADD COLUMN {col} {ddl}")
                        except sqlite3.OperationalError:
                            pass
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS saldos_mensuales (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        num_cuenta TEXT NOT NULL,
                        mes INTEGER NOT NULL,
                        anio INTEGER NOT NULL,
                        saldo_inicial REAL DEFAULT 0,
                        debe REAL DEFAULT 0,
                        haber REAL DEFAULT 0,
                        saldo_final REAL DEFAULT 0,
                        UNIQUE(num_cuenta, mes, anio)
                    );
                """)
                # Núcleo fase 1: partidas normalizadas, vínculo CFDI y saldos por periodo (motor).
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS partidas_poliza (
                        id_partida INTEGER PRIMARY KEY AUTOINCREMENT,
                        id_poliza INTEGER NOT NULL,
                        numero_linea INTEGER NOT NULL,
                        num_cuenta TEXT NOT NULL,
                        concepto_linea TEXT,
                        cargo REAL DEFAULT 0,
                        abono REAL DEFAULT 0,
                        cargo_mn REAL DEFAULT 0,
                        abono_mn REAL DEFAULT 0,
                        cliente_rfc TEXT,
                        cliente_nombre TEXT,
                        centro_costo_id INTEGER,
                        FOREIGN KEY(id_poliza) REFERENCES polizas(id) ON DELETE CASCADE
                    )
                """)
                # Migración: añadir centro a partidas si falta
                cur.execute("PRAGMA table_info(partidas_poliza)")
                pcols2 = [r[1] for r in cur.fetchall()]
                if "centro_costo_id" not in pcols2:
                    try:
                        cur.execute("ALTER TABLE partidas_poliza ADD COLUMN centro_costo_id INTEGER")
                    except sqlite3.OperationalError:
                        pass
                # Migración multimoneda: moneda/tipo_cambio por partida
                for col, ddl in (
                    ("moneda", "TEXT DEFAULT 'MXN'"),
                    ("tipo_cambio", "REAL DEFAULT 1.0"),
                ):
                    if col not in pcols2:
                        try:
                            cur.execute(f"ALTER TABLE partidas_poliza ADD COLUMN {col} {ddl}")
                        except sqlite3.OperationalError:
                            pass
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cfdi_poliza (
                        id_cfdi_poliza INTEGER PRIMARY KEY AUTOINCREMENT,
                        id_partida INTEGER NOT NULL,
                        uuid TEXT NOT NULL UNIQUE,
                        rfc_emisor TEXT,
                        rfc_receptor TEXT,
                        fecha_cfdi TEXT,
                        subtotal REAL,
                        iva_trasladado REAL,
                        iva_retenido REAL,
                        isr_retenido REAL,
                        total_cfdi REAL,
                        tipo_comprobante TEXT,
                        metodo_pago TEXT,
                        forma_pago TEXT,
                        xml_raw TEXT,
                        FOREIGN KEY(id_partida) REFERENCES partidas_poliza(id_partida) ON DELETE CASCADE
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS saldos_cuenta (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        num_cuenta TEXT NOT NULL,
                        ejercicio INTEGER NOT NULL,
                        periodo INTEGER NOT NULL,
                        saldo_inicial_mn REAL DEFAULT 0,
                        cargos_mn REAL DEFAULT 0,
                        abonos_mn REAL DEFAULT 0,
                        saldo_final_mn REAL DEFAULT 0,
                        UNIQUE(num_cuenta, ejercicio, periodo)
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS poliza_bitacora (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        poliza_id INTEGER NOT NULL,
                        accion TEXT NOT NULL,
                        detalle TEXT,
                        usuario TEXT,
                        creado_en TEXT NOT NULL,
                        FOREIGN KEY(poliza_id) REFERENCES polizas(id) ON DELETE CASCADE
                    )
                """)
                cur.execute("PRAGMA table_info(polizas)")
                _pcols_bit = [r[1] for r in cur.fetchall()]
                for _col, _ddl in (
                    ("usuario_afectacion", "TEXT"),
                    ("ts_afectacion", "TEXT"),
                    ("usuario_captura", "TEXT"),
                    ("alerta_cfdi", "TEXT"),
                ):
                    if _col not in _pcols_bit:
                        try:
                            cur.execute(f"ALTER TABLE polizas ADD COLUMN {_col} {_ddl}")
                        except sqlite3.OperationalError:
                            pass
                conn.commit()
        except Exception:
            # No romper la app si el esquema ya existe o hay permisos; los métodos reportarán error al operar.
            pass

    def _read_supervisor_polizas_password(self) -> str:
        env_p = os.environ.get("COI_SUPERVISOR_POLIZAS_PASSWORD", "").strip()
        if env_p:
            return env_p
        try:
            root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            cfg_path = os.path.join(root, "config_instituto.json")
            if os.path.isfile(cfg_path):
                with open(cfg_path, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
                return str(data.get("SUPERVISOR_POLIZAS_PASSWORD") or "").strip()
        except Exception:
            pass
        return ""

    @staticmethod
    def _sumas_mn_movimientos(movimientos: List[Dict[str, Any]], tipo_cambio_def: float = 1.0) -> Tuple[float, float]:
        """Suma cargos y abonos en moneda nacional (MN) para cuadrar pólizas multimoneda."""
        tc_d = float(tipo_cambio_def or 1.0)
        sc = sa = 0.0
        for m in movimientos:
            tc = float(m.get("tipo_cambio") or tc_d)
            cargo = float(m.get("cargo", 0) or 0)
            abono = float(m.get("abono", 0) or 0)
            if m.get("cargo_mn") is not None:
                cm = float(m["cargo_mn"])
            else:
                cm = cargo * tc
            if m.get("abono_mn") is not None:
                am = float(m["abono_mn"])
            else:
                am = abono * tc
            sc += cm
            sa += am
        return sc, sa

    def _registrar_bitacora_poliza(
        self,
        cur,
        poliza_id: int,
        accion: str,
        detalle: str = "",
        usuario: Optional[str] = None,
    ) -> None:
        try:
            cur.execute(
                """
                INSERT INTO poliza_bitacora (poliza_id, accion, detalle, usuario, creado_en)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    int(poliza_id),
                    str(accion or "")[:120],
                    (detalle or "")[:4000],
                    (usuario or "")[:120],
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
        except Exception:
            pass

    def _error_linea_credito_banco(self, fecha: str) -> Optional[str]:
        """
        Tras recalcular saldos: si una cuenta deudora tiene limite_credito_mxn > 0 y
        saldo_final del mes cae por debajo de -limite, se excede la línea/sobregiro.
        """
        try:
            fecha_obj = datetime.strptime((fecha or "")[:10], "%Y-%m-%d")
        except ValueError:
            return None
        mes, anio = fecha_obj.month, fecha_obj.year
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                try:
                    cur.execute("PRAGMA table_info(catalogo_cuentas)")
                    cols = [r[1] for r in cur.fetchall()]
                    if "limite_credito_mxn" not in cols:
                        return None
                except sqlite3.Error:
                    return None
                cur.execute(
                    """
                    SELECT num_cuenta, limite_credito_mxn, naturaleza
                    FROM catalogo_cuentas
                    WHERE limite_credito_mxn IS NOT NULL AND limite_credito_mxn > 0
                    """
                )
                for num, limite, nat in cur.fetchall():
                    limite = float(limite)
                    nat_u = str(nat or "").upper()
                    if nat_u != "DEUDORA":
                        continue
                    cur.execute(
                        """
                        SELECT saldo_final FROM saldos_mensuales
                        WHERE num_cuenta = ? AND mes = ? AND anio = ?
                        """,
                        (str(num).strip(), mes, anio),
                    )
                    row = cur.fetchone()
                    saldo = float(row[0]) if row and row[0] is not None else 0.0
                    if saldo < -limite - 0.02:
                        return (
                            f"Cuenta {num}: el saldo ({saldo:,.2f} MXN) excede la línea de "
                            f"crédito/sobregiro autorizada ({limite:,.2f} MXN). Ajuste movimientos o el límite en el catálogo."
                        )
        except Exception as e:
            return str(e)
        return None

    def _periodo_esta_bloqueado(self, anio: int, mes: int) -> bool:
        if PeriodosBloqueados is None:
            return False
        try:
            pb = PeriodosBloqueados(db_path=self.db_path)
            return bool(pb.esta_bloqueado(int(anio), int(mes)))
        except Exception:
            return False

    def _validar_periodo_abierto(self, fecha: str, operacion: str) -> Optional[str]:
        try:
            dt = datetime.strptime(str(fecha or "")[:10], "%Y-%m-%d")
        except Exception:
            return None
        if self._periodo_esta_bloqueado(dt.year, dt.month):
            return (
                f"Periodo {dt.month:02d}/{dt.year} bloqueado. "
                f"No se permite {operacion}."
            )
        return None

    def _folio_inicial_para_fecha(self, cursor: sqlite3.Cursor, tipo: str, fecha: str) -> int:
        """Folio mínimo configurado para el mes calendario (1 = sin override)."""
        try:
            cursor.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='folios_mensuales_config' LIMIT 1"
            )
            if not cursor.fetchone():
                return 1
            cursor.execute(
                """
                SELECT folio_inicial FROM folios_mensuales_config
                WHERE anio = CAST(strftime('%Y', ?) AS INTEGER)
                  AND mes = CAST(strftime('%m', ?) AS INTEGER)
                  AND tipo_poliza = ?
                LIMIT 1
                """,
                (fecha, fecha, str(tipo or "").strip().upper()),
            )
            row = cursor.fetchone()
            if row and row[0] is not None:
                return max(1, int(row[0]))
        except Exception:
            pass
        return 1

    def obtener_cabecera_poliza(self, poliza_id: int) -> Optional[Dict[str, Any]]:
        """Cabecera mínima (estatus, fecha, tipo, número) para validaciones de UI."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("PRAGMA table_info(polizas)")
                pcols = [r[1] for r in cur.fetchall()]
                sel = (
                    "id, numero_poliza, tipo_poliza, fecha, concepto, "
                    "UPPER(COALESCE(estatus,'C')) AS estatus"
                )
                if "usuario_captura" in pcols:
                    sel += ", usuario_captura, usuario_afectacion, ts_afectacion"
                if "alerta_cfdi" in pcols:
                    sel += ", alerta_cfdi"
                cur.execute(f"SELECT {sel} FROM polizas WHERE id = ?", (int(poliza_id),))
                row = cur.fetchone()
                return dict(row) if row else None
        except Exception:
            return None

    def _cta_valida_para_partida(self, cur, num_cuenta: str) -> Optional[str]:
        """None si la cuenta puede usarse en partida; mensaje de error en caso contrario."""
        num = str(num_cuenta or "").strip()
        if not num:
            return "Cuenta vacía en partida"
        try:
            cur.execute("PRAGMA table_info(catalogo_cuentas)")
            ccols = [r[1] for r in cur.fetchall()]
        except Exception:
            ccols = []
        if not ccols:
            return None
        cur.execute(
            "SELECT COALESCE(activa,1), COALESCE(UPPER(tipo_cuenta),'ACUMULATIVA') FROM catalogo_cuentas WHERE num_cuenta = ? LIMIT 1",
            (num,),
        )
        row = cur.fetchone()
        if not row:
            return f"Cuenta {num} no existe en el catálogo"
        if int(row[0] or 0) != 1:
            return f"Cuenta {num} está inactiva"
        if str(row[1] or "").strip().upper() != "DETALLE":
            return f"Cuenta {num} debe ser de tipo DETALLE para movimientos en pólizas"
        return None

    def _snapshot_poliza(self, poliza_id: int) -> Optional[Dict[str, Any]]:
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT tipo_poliza, fecha, concepto FROM polizas WHERE id = ?",
                    (int(poliza_id),),
                )
                row = cur.fetchone()
                if not row:
                    return None
                cur.execute(
                    "SELECT COUNT(*) FROM partidas_poliza WHERE id_poliza = ?",
                    (int(poliza_id),),
                )
                npp = int(cur.fetchone()[0] or 0)
                if npp > 0:
                    cur.execute(
                        """
                        SELECT num_cuenta, concepto_linea, cargo, abono, cliente_rfc, cliente_nombre, centro_costo_id
                        FROM partidas_poliza WHERE id_poliza = ? ORDER BY numero_linea, id_partida
                        """,
                        (int(poliza_id),),
                    )
                    movs = []
                    for r in cur.fetchall():
                        movs.append(
                            {
                                "num_cuenta": r[0],
                                "concepto": r[1] or "",
                                "cargo": float(r[2] or 0),
                                "abono": float(r[3] or 0),
                                "cliente_rfc": r[4],
                                "cliente_nombre": r[5],
                                "centro_costo_id": r[6],
                            }
                        )
                else:
                    cur.execute(
                        """
                        SELECT num_cuenta, concepto_mov, cargo, abono, cliente_rfc, cliente_nombre, centro_costo_id
                        FROM movimientos WHERE poliza_id = ? ORDER BY COALESCE(numero_linea, id), id
                        """,
                        (int(poliza_id),),
                    )
                    movs = []
                    for r in cur.fetchall():
                        movs.append(
                            {
                                "num_cuenta": r[0],
                                "concepto": r[1] or "",
                                "cargo": float(r[2] or 0),
                                "abono": float(r[3] or 0),
                                "cliente_rfc": r[4],
                                "cliente_nombre": r[5],
                                "centro_costo_id": r[6],
                            }
                        )
                return {
                    "tipo": row[0],
                    "fecha": row[1],
                    "concepto": row[2],
                    "movimientos": movs,
                }
        except Exception:
            return None

    def _restaurar_poliza(self, poliza_id: int, snap: Dict[str, Any]) -> None:
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute(
                "UPDATE polizas SET tipo_poliza = ?, fecha = ?, concepto = ? WHERE id = ?",
                (snap["tipo"], snap["fecha"], snap["concepto"], int(poliza_id)),
            )
            cur.execute("DELETE FROM movimientos WHERE poliza_id = ?", (int(poliza_id),))
            for idx, mov in enumerate(snap.get("movimientos") or [], start=1):
                cur.execute(
                    """
                    INSERT INTO movimientos (poliza_id, num_cuenta, concepto_mov, cargo, abono, cliente_rfc, cliente_nombre, centro_costo_id, numero_linea)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        int(poliza_id),
                        mov["num_cuenta"],
                        mov.get("concepto", mov.get("concepto_mov", "")),
                        float(mov.get("cargo", 0) or 0),
                        float(mov.get("abono", 0) or 0),
                        mov.get("cliente_rfc"),
                        mov.get("cliente_nombre"),
                        mov.get("centro_costo_id"),
                        int(idx),
                    ),
                )
            conn.commit()
    
    def crear_poliza(
        self,
        tipo: str,
        fecha: str,
        concepto: str,
        movimientos: List[Dict[str, Any]],
        *,
        moneda: str = "MXN",
        tipo_cambio: float = 1.0,
        estatus: str = "C",
        usuario_captura: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Crea una nueva póliza con sus movimientos
        movimientos: [
            {'num_cuenta': '101', 'concepto': '...', 'cargo': 1000, 'abono': 0},
            {'num_cuenta': '201', 'concepto': '...', 'cargo': 0, 'abono': 1000}
        ]
        """
        err_periodo = self._validar_periodo_abierto(fecha, "crear pólizas")
        if err_periodo:
            return {"exito": False, "error": err_periodo}

        total_cargos_mn, total_abonos_mn = self._sumas_mn_movimientos(movimientos, tipo_cambio)
        if abs(total_cargos_mn - total_abonos_mn) > 0.01:
            return {
                "exito": False,
                "error": (
                    f"Los cargos ({total_cargos_mn:,.2f} MN) deben ser iguales a los abonos "
                    f"({total_abonos_mn:,.2f} MN)"
                ),
            }
        
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                err_mon = self._validar_moneda_movimientos(cursor, movimientos)
                if err_mon:
                    return {"exito": False, "error": err_mon}
                for i, mov in enumerate(movimientos, start=1):
                    err_cta = self._cta_valida_para_partida(cursor, str(mov.get("num_cuenta") or "").strip())
                    if err_cta:
                        return {"exito": False, "error": f"Línea {i}: {err_cta}"}
                
                # Obtener el siguiente número de póliza para el tipo y fecha
                año_mes = fecha[:7]
                cursor.execute("""
                    SELECT COALESCE(MAX(numero_poliza), 0) + 1
                    FROM polizas
                    WHERE tipo_poliza = ? AND strftime('%Y-%m', fecha) = ?
                """, (tipo, año_mes))
                numero_poliza = int(cursor.fetchone()[0] or 0)
                base_folio = self._folio_inicial_para_fecha(cursor, tipo, fecha)
                numero_poliza = max(numero_poliza, base_folio)
                
                cursor.execute("PRAGMA table_info(polizas)")
                _ins_cols = [r[1] for r in cursor.fetchall()]
                usr_cap = ((usuario_captura or "").strip() or None)
                if "usuario_captura" in _ins_cols:
                    cursor.execute(
                        """
                        INSERT INTO polizas (numero_poliza, tipo_poliza, fecha, concepto, moneda, tipo_cambio, estatus, usuario_captura)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            numero_poliza,
                            tipo,
                            fecha,
                            concepto,
                            (moneda or "MXN").upper(),
                            float(tipo_cambio or 1.0),
                            (estatus or "C").upper()[:1],
                            usr_cap,
                        ),
                    )
                else:
                    cursor.execute(
                        """
                        INSERT INTO polizas (numero_poliza, tipo_poliza, fecha, concepto, moneda, tipo_cambio, estatus)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            numero_poliza,
                            tipo,
                            fecha,
                            concepto,
                            (moneda or "MXN").upper(),
                            float(tipo_cambio or 1.0),
                            (estatus or "C").upper()[:1],
                        ),
                    )
                
                poliza_id = cursor.lastrowid
                
                # Insertar movimientos (concepto_mov acepta 'concepto' del dict; opcional cliente_rfc/cliente_nombre)
                for idx, mov in enumerate(movimientos, start=1):
                    concepto = mov.get('concepto', mov.get('concepto_mov', ''))
                    tc_mov = float(mov.get("tipo_cambio") or tipo_cambio or 1.0)
                    cargo = float(mov.get('cargo', 0) or 0)
                    abono = float(mov.get('abono', 0) or 0)
                    cargo_mn = float(mov.get("cargo_mn") or (cargo * tc_mov))
                    abono_mn = float(mov.get("abono_mn") or (abono * tc_mov))
                    centro_id = mov.get("centro_costo_id", None)
                    cursor.execute("""
                        INSERT INTO movimientos (poliza_id, num_cuenta, concepto_mov, cargo, abono, cliente_rfc, cliente_nombre, centro_costo_id, numero_linea)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        poliza_id,
                        mov['num_cuenta'],
                        concepto,
                        cargo,
                        abono,
                        mov.get('cliente_rfc') or None,
                        mov.get('cliente_nombre') or None,
                        int(centro_id) if centro_id not in (None, "", 0) else None,
                        int(idx),
                    ))
                    cursor.execute(
                        """
                        INSERT INTO partidas_poliza
                        (id_poliza, numero_linea, num_cuenta, concepto_linea, cargo, abono, moneda, tipo_cambio, cargo_mn, abono_mn, cliente_rfc, cliente_nombre, centro_costo_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            poliza_id,
                            idx,
                            mov['num_cuenta'],
                            concepto,
                            cargo,
                            abono,
                            str(mov.get("moneda") or moneda or "MXN").upper(),
                            tc_mov,
                            cargo_mn,
                            abono_mn,
                            mov.get('cliente_rfc') or None,
                            mov.get('cliente_nombre') or None,
                            int(centro_id) if centro_id not in (None, "", 0) else None,
                        ),
                    )
                    id_partida = cursor.lastrowid
                    uuid = (mov.get("uuid") or "").strip()
                    if uuid:
                        exists = cursor.execute("SELECT 1 FROM cfdi_poliza WHERE uuid = ?", (uuid,)).fetchone()
                        if exists:
                            raise ValueError(f"UUID duplicado: {uuid}")
                        cursor.execute(
                            """
                            INSERT INTO cfdi_poliza
                            (id_partida, uuid, rfc_emisor, rfc_receptor, fecha_cfdi, subtotal, iva_trasladado, iva_retenido, isr_retenido, total_cfdi, tipo_comprobante, metodo_pago, forma_pago, xml_raw)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                id_partida,
                                uuid,
                                mov.get("rfc_emisor"),
                                mov.get("rfc_receptor"),
                                mov.get("fecha_cfdi"),
                                mov.get("subtotal"),
                                mov.get("iva_trasladado"),
                                mov.get("iva_retenido"),
                                mov.get("isr_retenido"),
                                mov.get("total_cfdi"),
                                mov.get("tipo_comprobante"),
                                mov.get("metodo_pago"),
                                mov.get("forma_pago"),
                                mov.get("xml_raw"),
                            ),
                        )
                
                conn.commit()

            self.recalcular_saldos_mensuales()
            err_lc = self._error_linea_credito_banco(fecha)
            if err_lc:
                try:
                    with sqlite3.connect(self.db_path) as conn:
                        c = conn.cursor()
                        c.execute("DELETE FROM movimientos WHERE poliza_id = ?", (poliza_id,))
                        c.execute("DELETE FROM polizas WHERE id = ?", (poliza_id,))
                        conn.commit()
                except Exception:
                    pass
                self.recalcular_saldos_mensuales()
                return {"exito": False, "error": err_lc}

            return {
                "exito": True,
                "numero_poliza": numero_poliza,
                "poliza_id": poliza_id,
                "tipo": tipo,
                "fecha": fecha,
            }
        except Exception as e:
            return {'exito': False, 'error': str(e)}
    
    def _naturaleza_cuenta(self, cursor, num_cuenta: str) -> str:
        """Devuelve 'DEUDORA' o 'ACREEDORA' según catalogo_cuentas."""
        try:
            cursor.execute(
                "SELECT naturaleza FROM catalogo_cuentas WHERE num_cuenta = ? LIMIT 1",
                (num_cuenta,)
            )
            row = cursor.fetchone()
            if row and row[0] in ('DEUDORA', 'ACREEDORA'):
                return row[0]
        except Exception:
            pass
        return 'DEUDORA'

    def _moneda_cuenta(self, cursor, num_cuenta: str) -> str:
        """Devuelve moneda configurada de la cuenta (default MXN)."""
        try:
            cursor.execute(
                "SELECT UPPER(COALESCE(moneda,'MXN')) FROM catalogo_cuentas WHERE num_cuenta = ? LIMIT 1",
                (str(num_cuenta or "").strip(),),
            )
            row = cursor.fetchone()
            if row and row[0]:
                return str(row[0]).strip().upper() or "MXN"
        except Exception:
            pass
        return "MXN"

    def _validar_moneda_movimientos(self, cursor, movimientos: List[Dict[str, Any]]) -> Optional[str]:
        """
        Regla multimoneda por cuenta:
        cada movimiento debe usar la misma moneda configurada en su cuenta contable.
        """
        for i, mov in enumerate(movimientos, start=1):
            cuenta = str(mov.get("num_cuenta") or "").strip()
            if not cuenta:
                return f"Movimiento #{i}: cuenta vacía."
            mon_mov = str(mov.get("moneda") or "MXN").strip().upper() or "MXN"
            mon_cta = self._moneda_cuenta(cursor, cuenta)
            if mon_mov != mon_cta:
                return (
                    f"Movimiento #{i} ({cuenta}): moneda {mon_mov} no coincide con "
                    f"moneda de cuenta ({mon_cta})."
                )
        return None

    def _actualizar_saldos(self, fecha: str, movimientos: List[Dict]):
        """Actualiza la tabla de saldos mensuales (DEUDORA: saldo = inicial + cargo - abono; ACREEDORA: saldo = inicial + abono - cargo)."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
                mes = fecha_obj.month
                anio = fecha_obj.year
                
                for mov in movimientos:
                    num_cuenta = mov['num_cuenta']
                    naturaleza = self._naturaleza_cuenta(cursor, num_cuenta)
                    cargo = mov.get('cargo', 0)
                    abono = mov.get('abono', 0)
                    
                    cursor.execute("""
                        SELECT id, saldo_inicial, debe, haber 
                        FROM saldos_mensuales 
                        WHERE num_cuenta = ? AND mes = ? AND anio = ?
                    """, (num_cuenta, mes, anio))
                    
                    registro = cursor.fetchone()
                    
                    if registro:
                        nuevo_debe = registro[2] + cargo
                        nuevo_haber = registro[3] + abono
                        if naturaleza == 'ACREEDORA':
                            nuevo_saldo = registro[1] + nuevo_haber - nuevo_debe
                        else:
                            nuevo_saldo = registro[1] + nuevo_debe - nuevo_haber
                        
                        cursor.execute("""
                            UPDATE saldos_mensuales 
                            SET debe = ?, haber = ?, saldo_final = ?
                            WHERE id = ?
                        """, (nuevo_debe, nuevo_haber, nuevo_saldo, registro[0]))
                    else:
                        if mes <= 1:
                            prev_anio, prev_mes = anio - 1, 12
                        else:
                            prev_anio, prev_mes = anio, mes - 1
                        cursor.execute("""
                            SELECT saldo_final FROM saldos_mensuales
                            WHERE num_cuenta = ? AND anio = ? AND mes = ?
                            ORDER BY anio DESC, mes DESC LIMIT 1
                        """, (num_cuenta, prev_anio, prev_mes))
                        
                        saldo_anterior = cursor.fetchone()
                        saldo_inicial = saldo_anterior[0] if saldo_anterior else 0
                        if naturaleza == 'ACREEDORA':
                            saldo_final = saldo_inicial + abono - cargo
                        else:
                            saldo_final = saldo_inicial + cargo - abono
                        
                        cursor.execute("""
                            INSERT INTO saldos_mensuales 
                            (num_cuenta, mes, anio, saldo_inicial, debe, haber, saldo_final)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (num_cuenta, mes, anio, saldo_inicial, cargo, abono, saldo_final))
        except Exception as e:
            print(f"Error actualizando saldos: {e}")
    
    def obtener_polizas(self, fecha_inicio: str, fecha_fin: str) -> List[Dict]:
        """Obtiene todas las pólizas en un rango de fechas"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT p.*,
                           GROUP_CONCAT(m.num_cuenta || ':' || m.cargo || ':' || m.abono) as movimientos,
                           GROUP_CONCAT(DISTINCT m.num_cuenta) as cuentas_list,
                           COALESCE(SUM(m.cargo), 0) as total_cargos,
                           COALESCE(SUM(m.abono), 0) as total_abonos
                    FROM polizas p
                    LEFT JOIN movimientos m ON p.id = m.poliza_id
                    WHERE p.fecha BETWEEN ? AND ?
                    GROUP BY p.id
                    ORDER BY p.fecha, p.numero_poliza
                """, (fecha_inicio, fecha_fin))
                
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error obteniendo pólizas: {e}")
            return []

    def buscar_polizas_avanzado(
        self,
        fecha_inicio: str,
        fecha_fin: str,
        *,
        tipo_poliza: Optional[str] = None,
        folio_desde: Optional[int] = None,
        folio_hasta: Optional[int] = None,
        num_cuenta: Optional[str] = None,
        concepto_like: Optional[str] = None,
        monto_aprox: Optional[float] = None,
        estatus: Optional[str] = None,
        usuario_captura_like: Optional[str] = None,
        usuario_afectacion_like: Optional[str] = None,
        solo_con_uuid: Optional[bool] = None,
    ) -> List[Dict[str, Any]]:
        """
        Filtros combinables sobre rango de fechas.
        solo_con_uuid: True = con CFDI ligado; False = sin CFDI; None = sin filtrar.
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("PRAGMA table_info(polizas)")
                pcols = [r[1] for r in cur.fetchall()]
                wheres = ["p.fecha BETWEEN ? AND ?"]
                params: List[Any] = [fecha_inicio, fecha_fin]
                tp = (tipo_poliza or "").strip().upper()
                if tp and tp != "TODAS":
                    wheres.append("UPPER(TRIM(p.tipo_poliza)) = ?")
                    params.append(tp)
                es = (estatus or "").strip().upper()
                if es and es != "TODOS":
                    wheres.append("UPPER(SUBSTR(COALESCE(p.estatus,'C'),1,1)) = ?")
                    params.append(es[:1])
                if folio_desde is not None:
                    wheres.append("p.numero_poliza >= ?")
                    params.append(int(folio_desde))
                if folio_hasta is not None:
                    wheres.append("p.numero_poliza <= ?")
                    params.append(int(folio_hasta))
                nc = (num_cuenta or "").strip()
                if nc:
                    like = f"%{nc}%"
                    wheres.append(
                        """(
                        EXISTS (SELECT 1 FROM movimientos mx WHERE mx.poliza_id = p.id AND mx.num_cuenta LIKE ?)
                        OR EXISTS (SELECT 1 FROM partidas_poliza px WHERE px.id_poliza = p.id AND px.num_cuenta LIKE ?)
                    )"""
                    )
                    params.extend([like, like])
                cl = (concepto_like or "").strip()
                if cl:
                    like = f"%{cl}%"
                    wheres.append(
                        """(
                        p.concepto LIKE ?
                        OR EXISTS (SELECT 1 FROM partidas_poliza pc WHERE pc.id_poliza = p.id AND COALESCE(pc.concepto_linea,'') LIKE ?)
                        OR EXISTS (SELECT 1 FROM movimientos mm WHERE mm.poliza_id = p.id AND COALESCE(mm.concepto_mov,'') LIKE ?)
                    )"""
                    )
                    params.extend([like, like, like])
                if monto_aprox is not None:
                    tol = 1.0
                    ma = float(monto_aprox)
                    wheres.append(
                        """(
                        EXISTS (
                          SELECT 1 FROM partidas_poliza pm WHERE pm.id_poliza = p.id
                          AND (
                            ABS(COALESCE(pm.cargo_mn, pm.cargo * COALESCE(pm.tipo_cambio,1.0)) - ?) <= ?
                            OR ABS(COALESCE(pm.abono_mn, pm.abono * COALESCE(pm.tipo_cambio,1.0)) - ?) <= ?
                          )
                        )
                        OR EXISTS (
                          SELECT 1 FROM movimientos mm2 WHERE mm2.poliza_id = p.id
                          AND (
                            ABS(COALESCE(mm2.cargo,0) - ?) <= ?
                            OR ABS(COALESCE(mm2.abono,0) - ?) <= ?
                          )
                        )
                    )"""
                    )
                    params.extend([ma, tol, ma, tol, ma, tol, ma, tol])
                uc = (usuario_captura_like or "").strip()
                if uc and "usuario_captura" in pcols:
                    wheres.append("COALESCE(p.usuario_captura,'') LIKE ?")
                    params.append(f"%{uc}%")
                ua = (usuario_afectacion_like or "").strip()
                if ua and "usuario_afectacion" in pcols:
                    wheres.append("COALESCE(p.usuario_afectacion,'') LIKE ?")
                    params.append(f"%{ua}%")
                if solo_con_uuid is True:
                    wheres.append(
                        """EXISTS (
                        SELECT 1 FROM partidas_poliza pu
                        INNER JOIN cfdi_poliza cf ON cf.id_partida = pu.id_partida
                        WHERE pu.id_poliza = p.id
                    )"""
                    )
                elif solo_con_uuid is False:
                    wheres.append(
                        """NOT EXISTS (
                        SELECT 1 FROM partidas_poliza pu
                        INNER JOIN cfdi_poliza cf ON cf.id_partida = pu.id_partida
                        WHERE pu.id_poliza = p.id
                    )"""
                    )
                sql = f"""
                    SELECT p.*,
                           GROUP_CONCAT(m.num_cuenta || ':' || m.cargo || ':' || m.abono) as movimientos,
                           GROUP_CONCAT(DISTINCT m.num_cuenta) as cuentas_list,
                           COALESCE(SUM(m.cargo), 0) as total_cargos,
                           COALESCE(SUM(m.abono), 0) as total_abonos
                    FROM polizas p
                    LEFT JOIN movimientos m ON p.id = m.poliza_id
                    WHERE {' AND '.join(wheres)}
                    GROUP BY p.id
                    ORDER BY p.fecha, p.tipo_poliza, p.numero_poliza
                """
                cur.execute(sql, params)
                return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            print(f"Error búsqueda avanzada pólizas: {e}")
            return []

    def consulta_poliza_completa(self, poliza_id: int, *, incluir_xml_raw: bool = True) -> Dict[str, Any]:
        """Cabecera, partidas y CFDI por partida (UUID, metadatos, xml opcional)."""
        cab = self.obtener_cabecera_poliza(int(poliza_id))
        if not cab:
            return {"exito": False, "error": "La póliza no existe"}
        try:
            partidas_out: List[Dict[str, Any]] = []
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT * FROM partidas_poliza WHERE id_poliza = ?
                    ORDER BY numero_linea, id_partida
                    """,
                    (int(poliza_id),),
                )
                for pr in cur.fetchall():
                    row = dict(pr)
                    pid = int(row["id_partida"])
                    cur.execute(
                        """
                        SELECT id_cfdi_poliza, uuid, rfc_emisor, rfc_receptor, fecha_cfdi, subtotal,
                               iva_trasladado, iva_retenido, isr_retenido, total_cfdi, tipo_comprobante,
                               metodo_pago, forma_pago
                               """
                        + (", xml_raw" if incluir_xml_raw else "")
                        + " FROM cfdi_poliza WHERE id_partida = ? ORDER BY id_cfdi_poliza",
                        (pid,),
                    )
                    cfdis = []
                    for cr in cur.fetchall():
                        d = dict(cr)
                        if not incluir_xml_raw:
                            d["xml_raw"] = None
                        cfdis.append(d)
                    row["cfdis"] = cfdis
                    partidas_out.append(row)
            return {"exito": True, "cabecera": cab, "partidas": partidas_out}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def importar_polizas_desde_json(
        self,
        data: Dict[str, Any],
        *,
        usuario_captura: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        data = { "polizas": [ { "tipo", "fecha", "concepto", "movimientos": [ {...} ] }, ... ] }
        Cada póliza queda en estatus C. Devuelve lista de errores por índice.
        """
        items = data.get("polizas") if isinstance(data, dict) else None
        if not isinstance(items, list):
            return {"exito": False, "error": "JSON inválido: se espera objeto con clave 'polizas' (arreglo)."}
        creadas: List[Dict[str, Any]] = []
        errores: List[Dict[str, Any]] = []
        for idx, bloque in enumerate(items):
            if not isinstance(bloque, dict):
                errores.append({"indice": idx, "error": "Elemento no es objeto"})
                continue
            tipo = str(bloque.get("tipo") or bloque.get("tipo_poliza") or "DIARIO").strip()
            fecha = str(bloque.get("fecha") or "").strip()[:10]
            concepto = str(bloque.get("concepto") or "").strip()
            movs = bloque.get("movimientos") or bloque.get("partidas")
            if not fecha or not concepto or not isinstance(movs, list) or len(movs) < 2:
                errores.append({"indice": idx, "error": "Faltan fecha, concepto o menos de 2 movimientos"})
                continue
            r = self.crear_poliza(
                tipo,
                fecha,
                concepto,
                list(movs),
                moneda=str(bloque.get("moneda") or "MXN"),
                tipo_cambio=float(bloque.get("tipo_cambio") or 1.0),
                estatus="C",
                usuario_captura=(usuario_captura or "").strip() or None,
            )
            if r.get("exito"):
                creadas.append({"indice": idx, "poliza_id": r.get("poliza_id"), "numero_poliza": r.get("numero_poliza")})
            else:
                errores.append({"indice": idx, "error": str(r.get("error") or "Error")})
        return {"exito": True, "creadas": creadas, "errores": errores, "total_creadas": len(creadas), "total_errores": len(errores)}

    def importar_polizas_desde_excel(self, path: str, *, usuario_captura: Optional[str] = None) -> Dict[str, Any]:
        """
        Hoja activa, encabezados: id_grupo, tipo, fecha, concepto, cuenta, concepto_mov, cargo, abono, moneda, tipo_cambio
        (id_grupo mismo número = misma póliza; moneda/tc opcionales por línea).
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"exito": False, "error": "Instale openpyxl."}
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            return {"exito": False, "error": str(e)}
        if not rows:
            return {"exito": False, "error": "Archivo vacío"}
        header = [str(c or "").strip().lower() for c in rows[0]]
        def col(name: str) -> Optional[int]:
            try:
                return header.index(name)
            except ValueError:
                return None
        ig = col("id_grupo")
        ti = col("tipo")
        fe = col("fecha")
        co = col("concepto")
        cu = col("cuenta")
        cm = col("concepto_mov")
        cg = col("cargo")
        ab = col("abono")
        mo = col("moneda")
        tc = col("tipo_cambio")
        if None in (ig, ti, fe, co, cu, cm, cg, ab):
            return {"exito": False, "error": "Faltan columnas requeridas: id_grupo, tipo, fecha, concepto, cuenta, concepto_mov, cargo, abono"}
        grupos: Dict[Any, Dict[str, Any]] = {}
        for ridx, r in enumerate(rows[1:], start=2):
            if not r or all(x is None or str(x).strip() == "" for x in r):
                continue
            try:
                g = r[ig]
                if g is None:
                    continue
                tipo = str(r[ti] or "DIARIO").strip()
                fecha_c = r[fe]
                if hasattr(fecha_c, "strftime"):
                    fecha_s = fecha_c.strftime("%Y-%m-%d")
                else:
                    fecha_s = str(fecha_c or "")[:10]
                concepto = str(r[co] or "").strip()
                cuenta = str(r[cu] or "").strip()
                conc_mov = str(r[cm] or "").strip()
                cargo = float(r[cg] or 0)
                abono = float(r[ab] or 0)
                mon = str(r[mo] if mo is not None and mo < len(r) and r[mo] else "MXN").upper() or "MXN"
                tcf = 1.0
                if tc is not None and tc < len(r) and r[tc] is not None:
                    try:
                        tcf = float(r[tc])
                    except (TypeError, ValueError):
                        tcf = 1.0
            except Exception as e:
                return {"exito": False, "error": f"Fila {ridx}: {e}"}
            if g not in grupos:
                grupos[g] = {"tipo": tipo, "fecha": fecha_s, "concepto": concepto, "movimientos": []}
            mov = {
                "num_cuenta": cuenta,
                "concepto": conc_mov,
                "cargo": cargo,
                "abono": abono,
                "moneda": mon,
                "tipo_cambio": tcf if tcf > 0 else 1.0,
                "cargo_mn": round(cargo * (tcf if tcf > 0 else 1.0), 6),
                "abono_mn": round(abono * (tcf if tcf > 0 else 1.0), 6),
            }
            grupos[g]["movimientos"].append(mov)
        payload = {"polizas": [grupos[k] for k in sorted(grupos.keys(), key=lambda x: (str(type(x)), str(x)))]}
        return self.importar_polizas_desde_json(payload, usuario_captura=usuario_captura)

    def listado_polizas_periodo(self, anio: int, mes: int) -> List[Dict[str, Any]]:
        """Lista plana fecha/tipo/folio/concepto/estatus para reimpresión."""
        mes_s = f"{int(mes):02d}"
        anio_s = str(int(anio))
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT id, fecha, tipo_poliza, numero_poliza, concepto,
                           UPPER(COALESCE(estatus,'C')) AS estatus
                    FROM polizas
                    WHERE strftime('%Y', fecha) = ? AND strftime('%m', fecha) = ?
                    ORDER BY tipo_poliza, numero_poliza, fecha
                    """,
                    (anio_s, mes_s),
                )
                return [dict(x) for x in cur.fetchall()]
        except Exception:
            return []

    def verificar_poliza(self, poliza_id: int) -> Dict[str, Any]:
        """Valida partida doble en MN, cuentas DETALLE activas, mínimo 2 partidas; C→V."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT id, estatus, fecha FROM polizas WHERE id = ?", (int(poliza_id),))
                row = cur.fetchone()
                if not row:
                    return {"exito": False, "error": "La póliza no existe"}
                err_periodo = self._validar_periodo_abierto(row[2], "verificar pólizas")
                if err_periodo:
                    return {"exito": False, "error": err_periodo}
                st = str(row[1] or "C").upper()[:1]
                if st == "X":
                    return {"exito": False, "error": "La póliza está cancelada"}
                if st == "A":
                    return {"exito": False, "error": "La póliza ya está afectada"}
                if st == "V":
                    return {"exito": True, "mensaje": "La póliza ya está verificada"}
                cur.execute("SELECT COUNT(*) FROM partidas_poliza WHERE id_poliza = ?", (int(poliza_id),))
                n_part = int(cur.fetchone()[0] or 0)
                if n_part == 0:
                    cur.execute("SELECT COUNT(*) FROM movimientos WHERE poliza_id = ?", (int(poliza_id),))
                    n_mov = int(cur.fetchone()[0] or 0)
                    if n_mov < 2:
                        return {"exito": False, "error": "Se requieren al menos 2 partidas"}
                    cur.execute(
                        """
                        SELECT num_cuenta, COALESCE(cargo,0), COALESCE(abono,0), COALESCE(tipo_cambio,1.0)
                        FROM movimientos WHERE poliza_id = ? ORDER BY COALESCE(numero_linea, id), id
                        """,
                        (int(poliza_id),),
                    )
                    rows = cur.fetchall()
                    sc = sa = 0.0
                    for num_c, cg, ab, tc in rows:
                        tc = float(tc or 1.0)
                        cg_f = float(cg or 0)
                        ab_f = float(ab or 0)
                        sc += cg_f * tc
                        sa += ab_f * tc
                        err_cta = self._cta_valida_para_partida(cur, str(num_c or "").strip())
                        if err_cta:
                            return {"exito": False, "error": err_cta}
                    if abs(sc - sa) > 0.01:
                        return {"exito": False, "error": "Póliza descuadrada en moneda nacional (cargos MN ≠ abonos MN)"}
                else:
                    if n_part < 2:
                        return {"exito": False, "error": "Se requieren al menos 2 partidas"}
                    cur.execute(
                        """
                        SELECT COALESCE(SUM(
                            COALESCE(cargo_mn, cargo * COALESCE(tipo_cambio,1.0))
                        ),0), COALESCE(SUM(
                            COALESCE(abono_mn, abono * COALESCE(tipo_cambio,1.0))
                        ),0)
                        FROM partidas_poliza WHERE id_poliza = ?
                        """,
                        (int(poliza_id),),
                    )
                    c, a = cur.fetchone() or (0, 0)
                    if abs(float(c or 0) - float(a or 0)) > 0.01:
                        return {"exito": False, "error": "Póliza descuadrada en moneda nacional (cargos MN ≠ abonos MN)"}
                    cur.execute(
                        "SELECT num_cuenta FROM partidas_poliza WHERE id_poliza = ? ORDER BY numero_linea, id_partida",
                        (int(poliza_id),),
                    )
                    for (num_c,) in cur.fetchall():
                        err_cta = self._cta_valida_para_partida(cur, str(num_c or "").strip())
                        if err_cta:
                            return {"exito": False, "error": err_cta}
                cur.execute("UPDATE polizas SET estatus = 'V' WHERE id = ?", (int(poliza_id),))
                conn.commit()
            return {"exito": True, "mensaje": "Póliza verificada"}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def _periodo_de_poliza(self, cur, poliza_id: int) -> tuple[int, int]:
        cur.execute("SELECT fecha FROM polizas WHERE id = ?", (int(poliza_id),))
        row = cur.fetchone()
        if not row or not row[0]:
            raise ValueError("Fecha de póliza inválida")
        dt = datetime.strptime(str(row[0])[:10], "%Y-%m-%d")
        return dt.year, dt.month

    def _saldo_final_por_naturaleza(self, cur, num_cuenta: str, ejercicio: int, periodo: int, delta_c: float, delta_a: float) -> tuple[float, float, float]:
        cur.execute(
            """
            SELECT saldo_inicial_mn, cargos_mn, abonos_mn
            FROM saldos_cuenta WHERE num_cuenta = ? AND ejercicio = ? AND periodo = ?
            """,
            (num_cuenta, ejercicio, periodo),
        )
        row = cur.fetchone()
        if row:
            saldo_inicial, cargos_mn, abonos_mn = float(row[0] or 0), float(row[1] or 0), float(row[2] or 0)
        else:
            saldo_inicial, cargos_mn, abonos_mn = 0.0, 0.0, 0.0
        cargos_mn += delta_c
        abonos_mn += delta_a
        nat = self._naturaleza_cuenta(cur, num_cuenta)
        saldo_final = saldo_inicial + (abonos_mn - cargos_mn if nat == "ACREEDORA" else cargos_mn - abonos_mn)
        return cargos_mn, abonos_mn, saldo_final

    def _aplicar_delta_en_cuenta_y_padres(self, cur, num_cuenta: str, ejercicio: int, periodo: int, delta_c: float, delta_a: float) -> None:
        actual = (num_cuenta or "").strip()
        visited = set()
        while actual and actual not in visited:
            visited.add(actual)
            cargos_mn, abonos_mn, saldo_final = self._saldo_final_por_naturaleza(cur, actual, ejercicio, periodo, delta_c, delta_a)
            cur.execute(
                """
                INSERT INTO saldos_cuenta (num_cuenta, ejercicio, periodo, saldo_inicial_mn, cargos_mn, abonos_mn, saldo_final_mn)
                VALUES (?, ?, ?, 0, ?, ?, ?)
                ON CONFLICT(num_cuenta, ejercicio, periodo)
                DO UPDATE SET cargos_mn=excluded.cargos_mn, abonos_mn=excluded.abonos_mn, saldo_final_mn=excluded.saldo_final_mn
                """,
                (actual, ejercicio, periodo, cargos_mn, abonos_mn, saldo_final),
            )
            cur.execute("SELECT cuenta_mayor FROM catalogo_cuentas WHERE num_cuenta = ? LIMIT 1", (actual,))
            row = cur.fetchone()
            actual = (row[0] if row and row[0] else "").strip()

    def afectar_poliza(self, poliza_id: int, *, usuario_afectacion: Optional[str] = None) -> Dict[str, Any]:
        """Afecta póliza verificada y propaga saldos a cuentas padre."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("BEGIN")
                cur.execute("SELECT estatus FROM polizas WHERE id = ?", (int(poliza_id),))
                row = cur.fetchone()
                if not row:
                    raise ValueError("La póliza no existe")
                st = str(row[0] or "C").upper()
                if st == "A":
                    return {"exito": True, "mensaje": "La póliza ya está afectada"}
                if st != "V":
                    raise ValueError("Solo se pueden afectar pólizas verificadas")
                ejercicio, periodo = self._periodo_de_poliza(cur, int(poliza_id))
                if self._periodo_esta_bloqueado(ejercicio, periodo):
                    raise ValueError(f"Periodo {periodo:02d}/{ejercicio} bloqueado. No se permite afectar pólizas.")
                cur.execute(
                    """
                    SELECT num_cuenta, COALESCE(cargo_mn,cargo,0), COALESCE(abono_mn,abono,0)
                    FROM partidas_poliza WHERE id_poliza = ? ORDER BY numero_linea
                    """,
                    (int(poliza_id),),
                )
                partes = cur.fetchall()
                if not partes:
                    raise ValueError("La póliza no tiene partidas")
                tc = sum(float(p[1] or 0) for p in partes)
                ta = sum(float(p[2] or 0) for p in partes)
                if abs(tc - ta) > 0.01:
                    raise ValueError("Póliza descuadrada")
                for num_cuenta, c_mn, a_mn in partes:
                    self._aplicar_delta_en_cuenta_y_padres(cur, str(num_cuenta), ejercicio, periodo, float(c_mn or 0), float(a_mn or 0))
                cur.execute("PRAGMA table_info(polizas)")
                pcols = [r[1] for r in cur.fetchall()]
                ts = datetime.now().isoformat(timespec="seconds")
                usr = (usuario_afectacion or "").strip() or None
                if "usuario_afectacion" in pcols and "ts_afectacion" in pcols:
                    cur.execute(
                        "UPDATE polizas SET estatus = 'A', usuario_afectacion = ?, ts_afectacion = ? WHERE id = ?",
                        (usr, ts, int(poliza_id)),
                    )
                else:
                    cur.execute("UPDATE polizas SET estatus = 'A' WHERE id = ?", (int(poliza_id),))
                conn.commit()
            self.recalcular_saldos_mensuales()
            return {"exito": True, "mensaje": "Póliza afectada"}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def _reconstruir_saldos_cuenta_desde_afectadas(self):
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM saldos_cuenta")
            cur.execute("SELECT id FROM polizas WHERE UPPER(COALESCE(estatus,'C')) = 'A' ORDER BY fecha, numero_poliza")
            ids = [int(r[0]) for r in cur.fetchall()]
            for pid in ids:
                ejercicio, periodo = self._periodo_de_poliza(cur, pid)
                cur.execute(
                    "SELECT num_cuenta, COALESCE(cargo_mn,cargo,0), COALESCE(abono_mn,abono,0) FROM partidas_poliza WHERE id_poliza = ? ORDER BY numero_linea",
                    (pid,),
                )
                for num_cuenta, c_mn, a_mn in cur.fetchall():
                    self._aplicar_delta_en_cuenta_y_padres(cur, str(num_cuenta), ejercicio, periodo, float(c_mn or 0), float(a_mn or 0))
            conn.commit()

    def desafectar_poliza(
        self,
        poliza_id: int,
        *,
        supervisor_password: str = "",
        justificacion: str = "",
        usuario_operador: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Revierte afectación y reconstruye saldos. Opcional: contraseña supervisor (config/env) y bitácora."""
        try:
            just = (justificacion or "").strip()
            if len(just) < 3:
                return {"exito": False, "error": "Indique una justificación (mínimo 3 caracteres)."}
            cfg_pass = self._read_supervisor_polizas_password()
            if cfg_pass and (supervisor_password or "").strip() != cfg_pass:
                return {"exito": False, "error": "Contraseña de supervisor incorrecta o no proporcionada."}
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT estatus FROM polizas WHERE id = ?", (int(poliza_id),))
                row = cur.fetchone()
                if not row:
                    return {"exito": False, "error": "La póliza no existe"}
                if str(row[0] or "").upper() != "A":
                    return {"exito": False, "error": "La póliza no está afectada"}
                ejercicio, periodo = self._periodo_de_poliza(cur, int(poliza_id))
                if self._periodo_esta_bloqueado(ejercicio, periodo):
                    return {"exito": False, "error": f"Periodo {periodo:02d}/{ejercicio} bloqueado. No se permite desafectar pólizas."}
                cur.execute("PRAGMA table_info(polizas)")
                pcols = [r[1] for r in cur.fetchall()]
                if "usuario_afectacion" in pcols and "ts_afectacion" in pcols:
                    cur.execute(
                        "UPDATE polizas SET estatus = 'V', usuario_afectacion = NULL, ts_afectacion = NULL WHERE id = ?",
                        (int(poliza_id),),
                    )
                else:
                    cur.execute("UPDATE polizas SET estatus = 'V' WHERE id = ?", (int(poliza_id),))
                self._registrar_bitacora_poliza(
                    cur,
                    int(poliza_id),
                    "DESAFECTAR",
                    just,
                    usuario_operador,
                )
                conn.commit()
            self._reconstruir_saldos_cuenta_desde_afectadas()
            self.recalcular_saldos_mensuales()
            return {"exito": True, "mensaje": "Póliza desafectada"}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def cancelar_poliza(self, poliza_id: int, motivo: str = "", *, usuario_operador: Optional[str] = None) -> Dict[str, Any]:
        """
        Cancela póliza:
        - Afectada (A): crea póliza espejo (cargos/abonos invertidos), verifica, afecta; original → X.
        - Verificada (V) o Captura (C): solo marca X (sin movimiento en saldos).
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(
                    "SELECT id, estatus, fecha, tipo_poliza, numero_poliza, concepto FROM polizas WHERE id = ?",
                    (int(poliza_id),),
                )
                prow = cur.fetchone()
                if not prow:
                    return {"exito": False, "error": "La póliza no existe"}
                estatus = str(prow["estatus"] or "C").upper()[:1]
                fecha = str(prow["fecha"] or "")
                tipo_o = str(prow["tipo_poliza"] or "DIARIO")
                num_o = int(prow["numero_poliza"] or 0)
                conc_o = str(prow["concepto"] or "")
                err_periodo = self._validar_periodo_abierto(fecha, "cancelar pólizas")
                if err_periodo:
                    return {"exito": False, "error": err_periodo}
                if estatus == "X":
                    return {"exito": True, "mensaje": "La póliza ya está cancelada"}

            motivo_s = (motivo or "").strip()
            ref = f"Cancelación pól. #{num_o}"
            if motivo_s:
                ref = f"{ref} — {motivo_s}"
            ref = ref[:240]

            if estatus == "A":
                with sqlite3.connect(self.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cur = conn.cursor()
                    cur.execute(
                        """
                        SELECT num_cuenta, concepto_linea, cargo, abono, moneda, tipo_cambio,
                               cargo_mn, abono_mn, cliente_rfc, cliente_nombre, centro_costo_id
                        FROM partidas_poliza WHERE id_poliza = ? ORDER BY numero_linea, id_partida
                        """,
                        (int(poliza_id),),
                    )
                    part_rows = cur.fetchall()
                if not part_rows:
                    return {"exito": False, "error": "La póliza afectada no tiene partidas para generar el espejo."}
                movs_espejo: List[Dict[str, Any]] = []
                for r in part_rows:
                    tc = float(r["tipo_cambio"] or 1.0)
                    cg = float(r["cargo"] or 0)
                    ab = float(r["abono"] or 0)
                    cm = r["cargo_mn"]
                    am = r["abono_mn"]
                    if cm is None:
                        cm = cg * tc
                    else:
                        cm = float(cm)
                    if am is None:
                        am = ab * tc
                    else:
                        am = float(am)
                    line_conc = str(r["concepto_linea"] or "")
                    movs_espejo.append(
                        {
                            "num_cuenta": str(r["num_cuenta"] or "").strip(),
                            "concepto": f"{ref}: {line_conc}"[:500],
                            "cargo": ab,
                            "abono": cg,
                            "moneda": str(r["moneda"] or "MXN").upper(),
                            "tipo_cambio": tc,
                            "cargo_mn": am,
                            "abono_mn": cm,
                            "cliente_rfc": r["cliente_rfc"],
                            "cliente_nombre": r["cliente_nombre"],
                            "centro_costo_id": int(r["centro_costo_id"])
                            if str(r["centro_costo_id"] or "").strip().isdigit()
                            else None,
                        }
                    )
                r_new = self.crear_poliza(
                    tipo_o,
                    fecha,
                    f"{ref} | Orig: {conc_o[:180]}",
                    movs_espejo,
                    moneda="MXN",
                    tipo_cambio=1.0,
                    estatus="C",
                    usuario_captura=(usuario_operador or "").strip() or "sistema",
                )
                if not r_new.get("exito"):
                    return r_new
                new_id = int(r_new["poliza_id"])
                rv = self.verificar_poliza(new_id)
                if not rv.get("exito"):
                    self.eliminar_poliza(new_id, permitir_cualquier_estatus=True)
                    return rv
                ra = self.afectar_poliza(new_id, usuario_afectacion=usuario_operador or "cancelación automática")
                if not ra.get("exito"):
                    self.eliminar_poliza(new_id, permitir_cualquier_estatus=True)
                    return ra
                with sqlite3.connect(self.db_path) as conn:
                    cur = conn.cursor()
                    cur.execute("UPDATE polizas SET estatus = 'X' WHERE id = ?", (int(poliza_id),))
                    self._registrar_bitacora_poliza(
                        cur,
                        int(poliza_id),
                        "CANCELAR",
                        f"Espejo póliza #{r_new.get('numero_poliza')} id={new_id}. {motivo_s}",
                        usuario_operador,
                    )
                    self._registrar_bitacora_poliza(
                        cur,
                        new_id,
                        "CANCELAR_ESPEJO",
                        f"Revierte póliza origen id={poliza_id} #{num_o}. {motivo_s}",
                        usuario_operador,
                    )
                    conn.commit()
                self.recalcular_saldos_mensuales()
                return {
                    "exito": True,
                    "mensaje": "Póliza cancelada con póliza espejo afectada",
                    "motivo": motivo_s,
                    "poliza_espejo_id": new_id,
                    "poliza_espejo_numero": r_new.get("numero_poliza"),
                }

            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("UPDATE polizas SET estatus = 'X' WHERE id = ?", (int(poliza_id),))
                self._registrar_bitacora_poliza(
                    cur,
                    int(poliza_id),
                    "CANCELAR",
                    motivo_s or "(sin motivo)",
                    usuario_operador,
                )
                conn.commit()

            return {"exito": True, "mensaje": "Póliza cancelada", "motivo": motivo_s}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def duplicar_poliza(
        self,
        poliza_id: int,
        *,
        fecha_nueva: Optional[str] = None,
        estatus_nuevo: str = "C",
        usuario_captura: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Duplica una póliza existente con nuevo consecutivo por tipo/mes."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(
                    "SELECT tipo_poliza, fecha, concepto, estatus FROM polizas WHERE id = ?",
                    (int(poliza_id),),
                )
                p = cur.fetchone()
                if not p:
                    return {"exito": False, "error": "La póliza no existe"}
                if str(p["estatus"] or "").upper()[:1] == "X":
                    return {"exito": False, "error": "No se puede duplicar una póliza cancelada."}
                fecha = str(fecha_nueva or p["fecha"] or "")[:10]
                err_periodo = self._validar_periodo_abierto(fecha, "duplicar pólizas")
                if err_periodo:
                    return {"exito": False, "error": err_periodo}
                cur.execute(
                    "SELECT COUNT(*) FROM partidas_poliza WHERE id_poliza = ?",
                    (int(poliza_id),),
                )
                n_part = int(cur.fetchone()[0] or 0)
                movs: List[Dict[str, Any]] = []
                if n_part > 0:
                    cur.execute(
                        """
                        SELECT num_cuenta, concepto_linea, cargo, abono, cliente_rfc, cliente_nombre,
                               centro_costo_id, moneda, tipo_cambio, cargo_mn, abono_mn
                        FROM partidas_poliza
                        WHERE id_poliza = ?
                        ORDER BY COALESCE(numero_linea, id_partida), id_partida
                        """,
                        (int(poliza_id),),
                    )
                    for r in cur.fetchall():
                        cid = r["centro_costo_id"]
                        tc = float(r["tipo_cambio"] or 1.0)
                        cargo = float(r["cargo"] or 0)
                        abono = float(r["abono"] or 0)
                        d: Dict[str, Any] = {
                            "num_cuenta": str(r["num_cuenta"] or "").strip(),
                            "concepto": str(r["concepto_linea"] or ""),
                            "cargo": cargo,
                            "abono": abono,
                            "cliente_rfc": r["cliente_rfc"],
                            "cliente_nombre": r["cliente_nombre"],
                            "moneda": str(r["moneda"] or "MXN").upper(),
                            "tipo_cambio": tc,
                            "cargo_mn": float(r["cargo_mn"] if r["cargo_mn"] is not None else (cargo * tc)),
                            "abono_mn": float(r["abono_mn"] if r["abono_mn"] is not None else (abono * tc)),
                        }
                        if cid not in (None, "", 0):
                            d["centro_costo_id"] = int(cid)
                        movs.append(d)
                else:
                    cur.execute(
                        """
                        SELECT num_cuenta, concepto_mov, cargo, abono, cliente_rfc, cliente_nombre, centro_costo_id
                        FROM movimientos
                        WHERE poliza_id = ?
                        ORDER BY COALESCE(numero_linea, id), id
                        """,
                        (int(poliza_id),),
                    )
                    for r in cur.fetchall():
                        cid = r["centro_costo_id"]
                        d: Dict[str, Any] = {
                            "num_cuenta": str(r["num_cuenta"] or "").strip(),
                            "concepto": str(r["concepto_mov"] or ""),
                            "cargo": float(r["cargo"] or 0),
                            "abono": float(r["abono"] or 0),
                            "cliente_rfc": r["cliente_rfc"],
                            "cliente_nombre": r["cliente_nombre"],
                        }
                        if cid not in (None, "", 0):
                            d["centro_costo_id"] = int(cid)
                        movs.append(d)

            if not movs:
                return {"exito": False, "error": "La póliza no tiene movimientos para duplicar"}

            r_new = self.crear_poliza(
                tipo=str(p["tipo_poliza"] or "DIARIO"),
                fecha=fecha,
                concepto=str(p["concepto"] or ""),
                movimientos=movs,
                moneda="MXN",
                tipo_cambio=1.0,
                estatus=(estatus_nuevo or "C").upper()[:1],
                usuario_captura=(usuario_captura or "").strip() or None,
            )
            if not r_new.get("exito"):
                return r_new
            return {
                "exito": True,
                "mensaje": "Póliza duplicada",
                "poliza_id": r_new.get("poliza_id"),
                "numero_poliza": r_new.get("numero_poliza"),
            }
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def vincular_cfdi_partida(self, id_partida: int, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Vincula UUID a partida (único global)."""
        uuid = str(payload.get("uuid") or "").strip()
        if not uuid:
            return {"exito": False, "error": "UUID obligatorio"}
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1 FROM partidas_poliza WHERE id_partida = ?", (int(id_partida),))
                if not cur.fetchone():
                    return {"exito": False, "error": "Partida no existe"}
                cur.execute("SELECT id_cfdi_poliza FROM cfdi_poliza WHERE uuid = ?", (uuid,))
                if cur.fetchone():
                    return {"exito": False, "error": f"UUID duplicado: {uuid}"}
                cur.execute(
                    """
                    INSERT INTO cfdi_poliza
                    (id_partida, uuid, rfc_emisor, rfc_receptor, fecha_cfdi, subtotal, iva_trasladado, iva_retenido, isr_retenido, total_cfdi, tipo_comprobante, metodo_pago, forma_pago, xml_raw)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        int(id_partida),
                        uuid,
                        payload.get("rfc_emisor"),
                        payload.get("rfc_receptor"),
                        payload.get("fecha_cfdi"),
                        payload.get("subtotal"),
                        payload.get("iva_trasladado"),
                        payload.get("iva_retenido"),
                        payload.get("isr_retenido"),
                        payload.get("total_cfdi"),
                        payload.get("tipo_comprobante"),
                        payload.get("metodo_pago"),
                        payload.get("forma_pago"),
                        payload.get("xml_raw"),
                    ),
                )
                conn.commit()
            return {"exito": True, "mensaje": "CFDI vinculado a partida"}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def desvincular_cfdi(self, id_cfdi_poliza: int) -> Dict[str, Any]:
        """Elimina el vínculo UUID–partida (el XML deja de estar ligado a esa línea)."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("DELETE FROM cfdi_poliza WHERE id_cfdi_poliza = ?", (int(id_cfdi_poliza),))
                if cur.rowcount == 0:
                    return {"exito": False, "error": "Registro CFDI no encontrado."}
                conn.commit()
            return {"exito": True, "mensaje": "CFDI desvinculado."}
        except Exception as ex:
            return {"exito": False, "error": str(ex)}

    def listar_cfdi_de_poliza(self, poliza_id: int) -> List[Dict[str, Any]]:
        """CFDIs ligados a partidas de una póliza (para desvincular o consultar)."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT c.id_cfdi_poliza, c.uuid, c.total_cfdi, c.fecha_cfdi, c.rfc_emisor, c.rfc_receptor,
                           pp.numero_linea, pp.num_cuenta
                    FROM cfdi_poliza c
                    JOIN partidas_poliza pp ON pp.id_partida = c.id_partida
                    WHERE pp.id_poliza = ?
                    ORDER BY pp.numero_linea
                    """,
                    (int(poliza_id),),
                )
                return [dict(r) for r in cur.fetchall()]
        except Exception:
            return []

    def actualizar_alerta_cfdi_poliza(self, poliza_id: int, texto: Optional[str]) -> None:
        """Marca texto de alerta por descuadre CFDI vs partidas (NULL limpia)."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("PRAGMA table_info(polizas)")
                if "alerta_cfdi" not in [r[1] for r in cur.fetchall()]:
                    return
                cur.execute(
                    "UPDATE polizas SET alerta_cfdi = ? WHERE id = ?",
                    ((texto.strip() if texto else None), int(poliza_id)),
                )
                conn.commit()
        except Exception:
            pass

    def editar_poliza(self, poliza_id: int, tipo: str, fecha: str, concepto: str,
                     movimientos: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Edita una póliza existente reemplazando completamente sus movimientos y
        recalculando saldos mensuales para mantener consistencia contable.
        Solo permitido en estatus Captura (C).
        """
        if len(movimientos) < 2:
            return {"exito": False, "error": "La póliza debe tener al menos 2 movimientos"}
        sc, sa = self._sumas_mn_movimientos(movimientos, 1.0)
        if abs(sc - sa) > 0.01:
            return {
                "exito": False,
                "error": f"Los cargos ({sc:,.2f} MN) deben ser iguales a los abonos ({sa:,.2f} MN)",
            }

        snap = self._snapshot_poliza(int(poliza_id))
        if not snap:
            return {'exito': False, 'error': 'La póliza no existe'}
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT UPPER(COALESCE(estatus,'C')) FROM polizas WHERE id = ?",
                    (int(poliza_id),),
                )
                er = cur.fetchone()
                if not er:
                    return {"exito": False, "error": "La póliza no existe"}
                if str(er[0] or "C").upper()[:1] != "C":
                    return {"exito": False, "error": "Solo se puede editar en estatus Captura (C). Verifique o desafecte antes."}
        except Exception:
            pass
        err_periodo_nuevo = self._validar_periodo_abierto(fecha, "editar pólizas")
        if err_periodo_nuevo:
            return {'exito': False, 'error': err_periodo_nuevo}
        err_periodo_original = self._validar_periodo_abierto(snap.get("fecha"), "editar pólizas")
        if err_periodo_original:
            return {'exito': False, 'error': err_periodo_original}

        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("PRAGMA foreign_keys = ON;")
                err_mon = self._validar_moneda_movimientos(cur, movimientos)
                if err_mon:
                    return {"exito": False, "error": err_mon}
                for i, mov in enumerate(movimientos, start=1):
                    err_cta = self._cta_valida_para_partida(cur, str(mov.get("num_cuenta") or "").strip())
                    if err_cta:
                        return {"exito": False, "error": f"Línea {i}: {err_cta}"}

                cur.execute("SELECT id FROM polizas WHERE id = ?", (int(poliza_id),))
                row = cur.fetchone()
                if not row:
                    return {'exito': False, 'error': 'La póliza no existe'}

                cur.execute("""
                    UPDATE polizas
                    SET tipo_poliza = ?, fecha = ?, concepto = ?
                    WHERE id = ?
                """, (tipo, fecha, concepto, int(poliza_id)))

                cur.execute("DELETE FROM movimientos WHERE poliza_id = ?", (int(poliza_id),))
                cur.execute("DELETE FROM partidas_poliza WHERE id_poliza = ?", (int(poliza_id),))
                for idx, mov in enumerate(movimientos, start=1):
                    concepto_mov = mov.get('concepto', mov.get('concepto_mov', ''))
                    tc_mov = float(mov.get("tipo_cambio") or 1.0)
                    cargo = float(mov.get('cargo', 0) or 0)
                    abono = float(mov.get('abono', 0) or 0)
                    cargo_mn = float(mov.get("cargo_mn") or (cargo * tc_mov))
                    abono_mn = float(mov.get("abono_mn") or (abono * tc_mov))
                    centro_id = mov.get("centro_costo_id", None)
                    cur.execute("""
                        INSERT INTO movimientos (poliza_id, num_cuenta, concepto_mov, cargo, abono, cliente_rfc, cliente_nombre, centro_costo_id, numero_linea)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        int(poliza_id),
                        str(mov.get('num_cuenta', '')).strip(),
                        concepto_mov,
                        cargo,
                        abono,
                        mov.get('cliente_rfc') or None,
                        mov.get('cliente_nombre') or None,
                        int(centro_id) if centro_id not in (None, "", 0) else None,
                        int(idx),
                    ))
                    # Re-crear partidas_poliza para que Verificar/Afectar trabajen con el mismo detalle
                    cur.execute(
                        """
                        INSERT INTO partidas_poliza
                        (id_poliza, numero_linea, num_cuenta, concepto_linea, cargo, abono, moneda, tipo_cambio, cargo_mn, abono_mn, cliente_rfc, cliente_nombre, centro_costo_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            int(poliza_id),
                            int(idx),
                            str(mov.get('num_cuenta', '')).strip(),
                            concepto_mov,
                            cargo,
                            abono,
                            str(mov.get("moneda") or "MXN").upper(),
                            tc_mov,
                            cargo_mn,
                            abono_mn,
                            mov.get('cliente_rfc') or None,
                            mov.get('cliente_nombre') or None,
                            int(centro_id) if centro_id not in (None, "", 0) else None,
                        ),
                    )
                conn.commit()

            self.recalcular_saldos_mensuales()
            err_lc = self._error_linea_credito_banco(fecha)
            if err_lc:
                self._restaurar_poliza(int(poliza_id), snap)
                self.recalcular_saldos_mensuales()
                return {"exito": False, "error": err_lc}
            return {'exito': True, 'poliza_id': int(poliza_id)}
        except Exception as e:
            try:
                self._restaurar_poliza(int(poliza_id), snap)
                self.recalcular_saldos_mensuales()
            except Exception:
                pass
            return {'exito': False, 'error': str(e)}

    def eliminar_poliza(self, poliza_id: int, *, permitir_cualquier_estatus: bool = False) -> Dict[str, Any]:
        """Elimina una póliza y sus movimientos. Por defecto solo en Captura (C); uso interno puede forzar."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("PRAGMA foreign_keys = ON;")
                cur.execute(
                    "SELECT id, fecha, UPPER(COALESCE(estatus,'C')) FROM polizas WHERE id = ?",
                    (int(poliza_id),),
                )
                row = cur.fetchone()
                if not row:
                    return {"exito": False, "error": "La póliza no existe"}
                if not permitir_cualquier_estatus and str(row[2] or "C").upper()[:1] != "C":
                    return {"exito": False, "error": "Solo se pueden eliminar pólizas en estatus Captura (C)."}
                err_periodo = self._validar_periodo_abierto(row[1], "eliminar pólizas")
                if err_periodo:
                    return {"exito": False, "error": err_periodo}
                cur.execute("DELETE FROM polizas WHERE id = ?", (int(poliza_id),))
                conn.commit()
            self.recalcular_saldos_mensuales()
            return {"exito": True, "poliza_id": int(poliza_id)}
        except Exception as e:
            return {"exito": False, "error": str(e)}

    def movimientos_por_poliza(self, poliza_id: int) -> List[Dict[str, Any]]:
        """
        Devuelve movimientos de una póliza específica, incluyendo cliente (si existe),
        para alimentar el resumen estilo COI.
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("""
                    SELECT
                        p.id as poliza_id,
                        p.fecha,
                        p.tipo_poliza,
                        p.numero_poliza,
                        p.concepto as concepto_poliza,
                        m.num_cuenta,
                        COALESCE(c.nombre_cuenta, '') as nombre_cuenta,
                        COALESCE(c.codigo_agrupador_sat, '') as tipo_sat,
                        COALESCE(c.naturaleza, '') as naturaleza,
                        COALESCE(c.tipo_cuenta, '') as tipo_cuenta,
                        m.concepto_mov,
                        m.cargo,
                        m.abono,
                        COALESCE(pp.moneda, 'MXN') as moneda,
                        COALESCE(pp.tipo_cambio, 1.0) as tipo_cambio,
                        COALESCE(pp.cargo_mn, (m.cargo * COALESCE(pp.tipo_cambio, 1.0))) as cargo_mn,
                        COALESCE(pp.abono_mn, (m.abono * COALESCE(pp.tipo_cambio, 1.0))) as abono_mn,
                        COALESCE(m.cliente_nombre, '') as cliente_nombre,
                        COALESCE(m.cliente_rfc, '') as cliente_rfc,
                        COALESCE(pp.centro_costo_id, m.centro_costo_id, NULL) as centro_costo_id
                    FROM movimientos m
                    JOIN polizas p ON m.poliza_id = p.id
                    LEFT JOIN catalogo_cuentas c ON c.num_cuenta = m.num_cuenta
                    LEFT JOIN partidas_poliza pp
                      ON pp.id_poliza = p.id
                     AND (
                       (m.numero_linea IS NOT NULL AND pp.numero_linea = m.numero_linea)
                       OR (
                         m.numero_linea IS NULL
                         AND pp.num_cuenta = m.num_cuenta
                         AND COALESCE(pp.concepto_linea,'') = COALESCE(m.concepto_mov,'')
                         AND COALESCE(pp.cargo,0) = COALESCE(m.cargo,0)
                         AND COALESCE(pp.abono,0) = COALESCE(m.abono,0)
                       )
                     )
                    WHERE p.id = ?
                    ORDER BY COALESCE(m.numero_linea, 999999), m.id ASC
                """, (int(poliza_id),))
                return [dict(r) for r in cur.fetchall()]
        except Exception:
            return []

    def exportar_polizas_periodo_excel(self, anio: int, mes: int, path: str) -> Dict[str, Any]:
        """Exporta pólizas del periodo con partidas a un archivo .xlsx (requiere openpyxl)."""
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"exito": False, "error": "Instale openpyxl para exportar a Excel."}
        mes_s = f"{int(mes):02d}"
        anio_s = str(int(anio))
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Polizas"
            headers = (
                "Fecha",
                "Tipo",
                "Folio",
                "Estatus",
                "Concepto poliza",
                "Linea",
                "Cuenta",
                "Concepto linea",
                "Cargo",
                "Abono",
                "Moneda",
                "TC",
                "Cargo MN",
                "Abono MN",
            )
            ws.append(headers)
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT p.fecha, p.tipo_poliza, p.numero_poliza,
                           COALESCE(p.estatus,'') AS estatus_pol,
                           p.concepto,
                           pp.numero_linea, pp.num_cuenta, pp.concepto_linea,
                           pp.cargo, pp.abono,
                           COALESCE(pp.moneda,'MXN') AS moneda_part,
                           COALESCE(pp.tipo_cambio,1.0) AS tc_part,
                           COALESCE(pp.cargo_mn, pp.cargo * COALESCE(pp.tipo_cambio,1.0)) AS cargo_mn_x,
                           COALESCE(pp.abono_mn, pp.abono * COALESCE(pp.tipo_cambio,1.0)) AS abono_mn_x
                    FROM polizas p
                    JOIN partidas_poliza pp ON pp.id_poliza = p.id
                    WHERE strftime('%Y', p.fecha) = ? AND strftime('%m', p.fecha) = ?
                    ORDER BY p.fecha, p.tipo_poliza, p.numero_poliza, pp.numero_linea
                    """,
                    (anio_s, mes_s),
                )
                for r in cur.fetchall():
                    ws.append(
                        [
                            r["fecha"],
                            r["tipo_poliza"],
                            r["numero_poliza"],
                            r["estatus_pol"],
                            r["concepto"],
                            r["numero_linea"],
                            r["num_cuenta"],
                            r["concepto_linea"],
                            float(r["cargo"] or 0),
                            float(r["abono"] or 0),
                            r["moneda_part"],
                            float(r["tc_part"] or 1.0),
                            float(r["cargo_mn_x"] or 0),
                            float(r["abono_mn_x"] or 0),
                        ]
                    )
            wb.save(path)
            return {"exito": True, "archivo": path}
        except Exception as e:
            return {"exito": False, "error": str(e)}

    def reporte_folios_periodo(self, tipo_poliza: str, anio: int, mes: int) -> Dict[str, Any]:
        """Folios del tipo y periodo; detecta duplicados y saltos en la correlatividad (excluye canceladas X)."""
        tipo = (tipo_poliza or "").strip().upper()
        mes_s = f"{int(mes):02d}"
        anio_s = str(int(anio))
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT numero_poliza FROM polizas
                    WHERE tipo_poliza = ?
                      AND strftime('%Y', fecha) = ?
                      AND strftime('%m', fecha) = ?
                      AND UPPER(COALESCE(estatus,'C')) <> 'X'
                    ORDER BY numero_poliza
                    """,
                    (tipo, anio_s, mes_s),
                )
                nums = [int(r[0]) for r in cur.fetchall() if r[0] is not None]
            cnt = Counter(nums)
            duplicados = sorted([n for n, k in cnt.items() if k > 1])
            faltantes: List[int] = []
            if nums:
                lo, hi = min(nums), max(nums)
                seen = set(nums)
                for n in range(lo, hi + 1):
                    if n not in seen:
                        faltantes.append(n)
            return {
                "exito": True,
                "tipo_poliza": tipo,
                "periodo": f"{anio_s}-{mes_s}",
                "folios": nums,
                "conteo": len(nums),
                "duplicados": duplicados,
                "faltantes": faltantes,
            }
        except Exception as e:
            return {"exito": False, "error": str(e)}

    def recalcular_saldos_mensuales(self) -> None:
        """Reconstruye saldos_mensuales desde polizas y movimientos, respetando naturaleza DEUDORA/ACREEDORA."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("PRAGMA table_info(polizas)")
                pcols = [r[1] for r in cursor.fetchall()]
                where_estatus = "WHERE UPPER(COALESCE(p.estatus,'A')) = 'A'" if "estatus" in pcols else ""
                cursor.execute(f"""
                    SELECT m.num_cuenta,
                           CAST(strftime('%m', p.fecha) AS INTEGER) as mes,
                           CAST(strftime('%Y', p.fecha) AS INTEGER) as anio,
                           SUM(m.cargo) as debe,
                           SUM(m.abono) as haber
                    FROM movimientos m
                    JOIN polizas p ON m.poliza_id = p.id
                    {where_estatus}
                    GROUP BY m.num_cuenta, mes, anio
                    ORDER BY m.num_cuenta, anio, mes
                """)
                filas = cursor.fetchall()
                if not filas:
                    cursor.execute("DELETE FROM saldos_mensuales")
                    conn.commit()
                    return
                last_saldo = {}
                for row in filas:
                    num_cuenta, mes, anio, debe, haber = row[0], row[1], row[2], float(row[3] or 0), float(row[4] or 0)
                    naturaleza = self._naturaleza_cuenta(cursor, num_cuenta)
                    saldo_inicial = last_saldo.get(num_cuenta, 0.0)
                    if naturaleza == 'ACREEDORA':
                        saldo_final = saldo_inicial + haber - debe
                    else:
                        saldo_final = saldo_inicial + debe - haber
                    last_saldo[num_cuenta] = saldo_final
                    cursor.execute("""
                        INSERT OR REPLACE INTO saldos_mensuales
                        (num_cuenta, mes, anio, saldo_inicial, debe, haber, saldo_final)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (num_cuenta, mes, anio, saldo_inicial, debe, haber, saldo_final))
                conn.commit()
        except Exception as e:
            print(f"Error recalculando saldos: {e}")